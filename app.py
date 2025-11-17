import atexit
import csv
import io
import json
import os
import shutil
import sys
import sqlite3
import uuid
from collections import defaultdict
from copy import deepcopy
from typing import Dict, List, Optional
from datetime import datetime, timedelta, timezone
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from pathlib import Path
from urllib.parse import urlparse
import socket
from waitress import serve
import random

import requests
import stat
from dateutil import tz, parser
from flask import (
    Flask,
    Response,
    flash,
    jsonify,
    redirect,
    render_template,
    render_template_string,
    request,
    send_file,
    session,
    url_for,
)
from flask_migrate import Migrate
from sqlalchemy import func, inspect, or_
from sqlalchemy.orm import joinedload

from analytics import (
    get_customer_retention,
    get_day_wise_billing,
    get_sales_trends,
    get_top_customers,
)
from analytics_tracking import log_user_event
from api import api_bp
from db import db_events  # noqa: F401
from db.models import db, customer, invoice, invoiceItem, item, layoutConfig, accountingTransaction, expenseItem
from migration import migrate_db
from supabase_upload import SupabaseUploadError, upload_full_database, upload_to_supabase

APP_NAME = "SLO BILL"
BG_DESKTOP_ENV = "BG_DESKTOP"
DEFAULT_SECRET_KEY = "super-secret"
DATABASE_FILENAME = "app.db"
INFO_FILENAME = "info.json"
APP_VERSION = "3.1.2"
DEFAULT_TIMEZONE = "Asia/Kolkata"
REQUIRED_DB_TABLES = {"customer", "invoice", "invoice_item", "item"}
BACKUP_DIRNAME = "backups"
BACKUP_RETENTION = 10
BACKUP_MAX_AGE_DAYS = 7
ISO_8601_UTC = "%Y-%m-%dT%H:%M:%SZ"
HUMAN_DATE_FMT = "%d %B %Y"
DEFAULT_LOGO_COLOR_MODE = "black"
LOGO_COLOR_PATHS = {
    "black": "static/img/brand-water-mark-black.svg",
    "blue": "static/img/brand-water-mark-blue.svg",
}
LOGO_COLOR_VALUES = {
    "black": "#111111",
    "blue": "#1b4ea0",
}
ACCOUNTING_STATEMENT_DEFAULT_START = datetime(2025, 9, 1).date()


def _default_info_sections(reference_dt: Optional[datetime] = None) -> dict:
    """Return a fresh copy of default info.json sections."""
    reference_dt = reference_dt or datetime.now(timezone.utc)
    iso_now = reference_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
    current_year = str(reference_dt.year)

    return {
        "business": {
            "name": "",
            "owner": "",
            "address": "",
            "email": "",
            "phone": "",
            "gstin": "",
            "upi_id": "",
            "upi_name": "",
            "businessType": "",
            "pan": "",
            "estd": current_year,
            "logo_path": "static/img/brand-wordmark.svg",
            "logo_color_mode": DEFAULT_LOGO_COLOR_MODE,
        },
        "bank": {
            "account_name": "",
            "bank_name": "",
            "branch": "",
            "account_number": "",
            "ifsc": "",
            "bhim": "",
        },
        "appearance": {
            "currency_symbol": "\\u20b9",
            "date_format": "%d %B %Y",
            "font_family": "Inter, Helvetica, Arial, sans-serif",
            "theme_color": "#0056b3",
        },
        "payment": {
            "methods": ["Bank Transfer", "UPI"],
            "upi_qr_enabled": False,
            "qr_label": "Scan to Pay",
            "terms": [
                "Payment due within 7 days of invoice date.",
                "Late payments may incur a 2% interest per month.",
                "This is a computer-generated document, no signature required.",
            ],
        },
        "statement": {
            "header_title": "Statement Summary",
            "disclaimer": "This is a system-generated statement. No signature required.",
        },
        "account_defaults": {
            "start_date": iso_now,
            "timezone": DEFAULT_TIMEZONE,
        },
        "meta": {
            "version": APP_VERSION,
            "created_on": iso_now,
        },
        "upi_info": {
            "upi_id": "",
            "currency": "INR",
            "upi_name": "",
        },
        "services": [
            "Printing",
            "Design",
            "Branding Collateral",
        ],
        "bill_config": {
            "heading": "Tax Invoice",
            "footer": "Composition Taxable Person. Not eligible to collect Tax on supplies.",
            "payment-footer": "Computer generated receipt - Signature not required",
        },
        "file_location": "",
        "supabase": {
            "url": "",
            "key": "",
            "last_uploaded": "",
            "last_incremental_uploaded": "",
        },
    }


def _merge_missing(target: dict, defaults: dict) -> bool:
    """Merge missing keys from defaults into target. Returns True if mutated."""
    changed = False
    for key, default_value in defaults.items():
        if key not in target:
            target[key] = deepcopy(default_value)
            changed = True
        else:
            current_value = target[key]
            if isinstance(default_value, dict) and isinstance(current_value, dict):
                if _merge_missing(current_value, default_value):
                    changed = True
    return changed


def _ensure_utc(dt: datetime) -> datetime:
    """Normalize datetimes to UTC with timezone info."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _format_iso_utc(dt: datetime) -> str:
    return _ensure_utc(dt).strftime(ISO_8601_UTC)


def _format_human_date(dt: datetime) -> str:
    return _ensure_utc(dt).strftime(HUMAN_DATE_FMT)


def _resolve_brand_watermark_path(business_section: Optional[dict]) -> str:
    """Return the static asset path for the selected brand watermark color."""
    if not isinstance(business_section, dict):
        business_section = {}
    color_mode = (business_section.get("logo_color_mode") or DEFAULT_LOGO_COLOR_MODE).lower()
    return LOGO_COLOR_PATHS.get(color_mode, LOGO_COLOR_PATHS[DEFAULT_LOGO_COLOR_MODE])


def _resolve_brand_accent_color(business_section: Optional[dict]) -> str:
    """Return the hex color used for accent text."""
    if not isinstance(business_section, dict):
        business_section = {}
    color_mode = (business_section.get("logo_color_mode") or DEFAULT_LOGO_COLOR_MODE).lower()
    return LOGO_COLOR_VALUES.get(color_mode, LOGO_COLOR_VALUES[DEFAULT_LOGO_COLOR_MODE])


def _get_earliest_invoice_created_at() -> Optional[datetime]:
    """Fetch the earliest invoice.createdAt value from the DB."""
    try:
        with app.app_context():
            earliest = (
                db.session.query(func.min(invoice.createdAt))
                .filter(invoice.isDeleted == False)  # noqa: E712
                .scalar()
            )
    except Exception as exc:
        print(f"[warn] Failed to determine earliest invoice date: {exc}")
        return None

    if not earliest:
        return None

    if isinstance(earliest, str):
        try:
            earliest = datetime.strptime(earliest, ISO_8601_UTC)
        except Exception:
            return None

    return _ensure_utc(earliest)


def _determine_data_start(now: datetime) -> datetime:
    """Return the correct account start date, preferring the first invoice date."""
    earliest = _get_earliest_invoice_created_at()
    if not earliest:
        return now

    earliest_utc = _ensure_utc(earliest)
    # Guard against corrupted future dates
    if earliest_utc > now:
        return now
    return earliest_utc


def _issue_bill_token() -> str:
    token = uuid.uuid4().hex
    session['bill_form_token'] = token
    return token


def _validate_bill_token(submitted: str) -> bool:
    expected = session.get('bill_form_token')
    if not expected or not submitted or submitted != expected:
        return False
    session.pop('bill_form_token', None)
    return True


def _render_create_bill(**context):
    context['form_token'] = _issue_bill_token()
    return render_template('create_bill.html', **context)


def _ensure_file_writable(path: Path) -> None:
    """Best-effort to guarantee the SQLite file is writable (fixes Windows bundle perms)."""
    try:
        if not path.exists():
            return
        if os.name == "nt":
            path.chmod(stat.S_IWRITE | stat.S_IREAD)
        else:
            path.chmod(0o600)
    except Exception as exc:
        print(f"[warn] Could not adjust permissions for {path}: {exc}")


def _desktop_data_dir(app_name: str) -> Path:
    if os.name == "nt":
        return Path(os.getenv("APPDATA", str(Path.home() / "AppData" / "Roaming"))) / app_name
    if sys.platform == "darwin":
        return Path.home() / "Library" / "Application Support" / app_name
    return Path.home() / ".local" / "share" / app_name


BASE_DIR = Path(__file__).resolve().parent
IS_DESKTOP = os.getenv(BG_DESKTOP_ENV) == "1"
DATA_DIR = _desktop_data_dir(APP_NAME) if IS_DESKTOP else BASE_DIR / "db"
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / DATABASE_FILENAME
SEED_DB_PATH = BASE_DIR / "db" / DATABASE_FILENAME
INFO_PATH = DATA_DIR / INFO_FILENAME


app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", DEFAULT_SECRET_KEY)
app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_PATH.as_posix()}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.register_blueprint(api_bp)

migrate_db(DB_PATH.as_posix())

# Attach to this Flask app
db.init_app(app)
migrate = Migrate(app, db)


def _format_customer_id(n: int) -> str:
    return f"ID-{n:06d}"


def rounding_to_nearest_zero(amount):
    """Rounding number to nearest zero"""
    try:
        d = Decimal(str(amount))
    except Exception:
        d = Decimal('0')
    tens = (d / Decimal('10')).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
    return float(tens * Decimal('10'))


def _ensure_db_initialized():
    """
    Ensure database schema exists.
    - If DB file missing: create schema (or copy seed if present + desktop).
    - If DB file exists but has no tables: create schema.
    """
    with app.app_context():
        # create file if missing (and optionally copy seed on desktop)
        if not DB_PATH.exists():
            try:
                if SEED_DB_PATH.exists() and IS_DESKTOP:
                    shutil.copy2(SEED_DB_PATH, DB_PATH)
                    print("[info] Copied seed DB to desktop data dir.")
                else:
                    # touch file so engine can open it cleanly
                    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
                    DB_PATH.touch(exist_ok=True)
                    print("[info] Created empty DB file.")
            except Exception as e:
                print(f"[warn] could not prepare DB file: {e}")
        # Always ensure the DB file is writable (pyinstaller seed copies can be read-only on Windows)
        _ensure_file_writable(DB_PATH)
        if DB_PATH.exists() and os.name == "nt":
            try:
                test_conn = sqlite3.connect(DB_PATH.as_posix())
                test_conn.execute("PRAGMA user_version;")
                test_conn.close()
            except sqlite3.OperationalError as exc:
                if "readonly" in str(exc).lower():
                    print("[warn] Detected read-only SQLite database; retrying permission reset.")
                    _ensure_file_writable(DB_PATH)
                else:
                    print(f"[warn] SQLite connection check failed: {exc}")

        # check whether tables exist
        insp = inspect(db.engine)
        tables = set(insp.get_table_names())

        if not REQUIRED_DB_TABLES.issubset(tables):
            print("[info] Creating/migrating schema via create_all()…")
            db.create_all()


def get_info_json_path():
    """Return correct info.json path"""
    return INFO_PATH


def ensure_info_json():
    """ensure db info.json exists or else creates it"""
    info_path = get_info_json_path()
    if not info_path.parent.exists():
        info_path.parent.mkdir(parents=True, exist_ok=True)

    now = datetime.now(timezone.utc)
    determined_start = _determine_data_start(now)
    start_iso = _format_iso_utc(determined_start)
    start_dt = _ensure_utc(datetime.strptime(start_iso, ISO_8601_UTC))
    created_display = _format_human_date(start_dt)

    default_sections = _default_info_sections(start_dt)
    default_sections["account_defaults"]["start_date"] = start_iso
    default_sections["meta"]["created_on"] = start_iso

    default_payload = {
        "created_on": created_display,
        "app_name": APP_NAME,
        "version": APP_VERSION,
        "last_updated": now.strftime(HUMAN_DATE_FMT),
        "onboarding_complete": False,
        "data": default_sections,
    }

    if not info_path.exists():
        try:
            with open(info_path, "w", encoding='utf-8') as f:
                json.dump(default_payload, f, indent=2, ensure_ascii=False)
            print("[info] Created info.json with default structure.")
        except Exception as e:
            print(f"[warn] could not create db info.json: {e}")
        return info_path

    try:
        with open(info_path, "r", encoding='utf-8') as f:
            info_data = json.load(f)
    except Exception as exc:
        print(f"[warn] Failed to read info.json ({exc}); rewriting with defaults.")
        try:
            with open(info_path, "w", encoding='utf-8') as f:
                json.dump(default_payload, f, indent=2, ensure_ascii=False)
        except Exception as write_err:
            print(f"[warn] could not rewrite db info.json: {write_err}")
        return info_path

    changed = False
    for key in ("created_on", "app_name", "version", "last_updated"):
        if key not in info_data:
            info_data[key] = default_payload[key]
            changed = True

    if "onboarding_complete" not in info_data:
        info_data["onboarding_complete"] = False
        changed = True

    if not isinstance(info_data.get("data"), dict):
        info_data["data"] = deepcopy(default_sections)
        changed = True
    else:
        if _merge_missing(info_data["data"], default_sections):
            changed = True

    data_section = info_data["data"]

    account_defaults = data_section.setdefault("account_defaults", {})
    existing_start = account_defaults.get("start_date")
    if existing_start:
        try:
            existing_start_dt = _ensure_utc(datetime.strptime(existing_start, ISO_8601_UTC))
        except Exception:
            account_defaults["start_date"] = start_iso
            changed = True
        else:
            if existing_start_dt > start_dt:
                account_defaults["start_date"] = start_iso
                changed = True
    else:
        account_defaults["start_date"] = start_iso
        changed = True

    meta_section = data_section.setdefault("meta", {})
    meta_created = meta_section.get("created_on")
    if meta_created:
        try:
            meta_dt = _ensure_utc(datetime.strptime(meta_created, ISO_8601_UTC))
        except Exception:
            meta_section["created_on"] = start_iso
            changed = True
        else:
            if meta_dt != start_dt:
                meta_section["created_on"] = start_iso
                changed = True
    else:
        meta_section["created_on"] = start_iso
        changed = True

    created_on_value = info_data.get("created_on")
    if created_on_value:
        try:
            created_on_dt = datetime.strptime(created_on_value, HUMAN_DATE_FMT).replace(tzinfo=timezone.utc)
        except Exception:
            info_data["created_on"] = created_display
            changed = True
        else:
            if created_on_dt.date() != start_dt.date():
                info_data["created_on"] = created_display
                changed = True
    else:
        info_data["created_on"] = created_display
        changed = True

    if changed:
        try:
            with open(info_path, "w", encoding='utf-8') as f:
                json.dump(info_data, f, indent=2, ensure_ascii=False)
        except Exception as exc:
            print(f"[warn] Failed to update info.json defaults: {exc}")

    return info_path


def loading_info():
    info_path = ensure_info_json()

    with open(info_path, 'r', encoding='utf-8') as f:
        json_loaded = json.load(f)
        return json_loaded


def refresh_info_json():
    """Reload the info.json without restarting the app"""
    global APP_INFO, ONBOARDING_COMPLETE
    try:
        full_payload = loading_info()
        new_info = full_payload.get('data', {})
        if not isinstance(new_info, dict):
            new_info = {}
        APP_INFO.clear()
        APP_INFO.update(new_info)
        ONBOARDING_COMPLETE = bool(full_payload.get('onboarding_complete', False))
    except Exception as e:
        print(f"[warn] Failed to load/refresh app_info: {e}")


_initial_info_payload = loading_info()
APP_INFO = _initial_info_payload.get('data', {})
if not isinstance(APP_INFO, dict):
    APP_INFO = {}
ONBOARDING_COMPLETE = bool(_initial_info_payload.get('onboarding_complete', False))


ONBOARDING_EXEMPT_ENDPOINTS = {
    'static',
    'onboarding',
    'onboarding_submit',
    'config_refresh',
}


def _is_onboarding_complete() -> bool:
    return bool(ONBOARDING_COMPLETE)


def _clean_analytics_payload(data: dict) -> dict:
    """Remove empty values and normalise strings from analytics payloads."""
    if not isinstance(data, dict):
        return {}

    cleaned = {}
    for key, value in data.items():
        if isinstance(value, str):
            value = value.strip()
        if value in (None, ""):
            continue
        cleaned[key] = value
    return cleaned


@app.before_request
def _enforce_onboarding_flow():
    if _is_onboarding_complete():
        return

    endpoint = request.endpoint or ''
    if endpoint in ONBOARDING_EXEMPT_ENDPOINTS:
        return
    if endpoint.startswith('api_bp.'):
        return

    return redirect(url_for('onboarding'))


@app.route('/onboarding', methods=['GET'])
def onboarding():
    if _is_onboarding_complete():
        return redirect(url_for('home'))

    seed_info = loading_info().get('data', {})
    business_defaults = seed_info.get('business', {}) if isinstance(seed_info, dict) else {}
    bank_defaults = seed_info.get('bank', {}) if isinstance(seed_info, dict) else {}
    return render_template(
        'onboarding.html',
        business=business_defaults,
        bank=bank_defaults,
    )


def _normalize_account_number(value: str) -> str:
    return ''.join(ch for ch in value if ch.isalnum())


def _get_upi_variants() -> List[Dict[str, str]]:
    """Return configured UPI variants with display metadata."""
    upi_info = APP_INFO.get('upi_info', {}) if isinstance(APP_INFO.get('upi_info'), dict) else {}
    business_info = APP_INFO.get('business', {}) if isinstance(APP_INFO.get('business'), dict) else {}

    def _clean(value: Optional[str]) -> Optional[str]:
        if not value:
            return None
        stripped = value.strip()
        return stripped or None

    variants: List[Dict[str, str]] = []

    primary_id = _clean(upi_info.get('upi_id') or business_info.get('upi_id'))
    primary_name = _clean(upi_info.get('upi_name') or business_info.get('upi_name') or business_info.get('owner'))
    if primary_id:
        variants.append({
            'key': 'primary',
            'label': 'Savings Account UPI',
            'upi_id': primary_id,
            'upi_name': primary_name or '',
        })

    current_id = _clean(upi_info.get('upi_current_id'))
    current_name = _clean(upi_info.get('upi_current_name') or primary_name)
    if current_id:
        variants.append({
            'key': 'current',
            'label': 'Current Account UPI',
            'upi_id': current_id,
            'upi_name': current_name or '',
        })

    return variants


def _find_upi_variant(choice: Optional[str]) -> Optional[Dict[str, str]]:
    if not choice:
        return None
    for variant in _get_upi_variants():
        if variant.get('key') == choice:
            return variant
    return None


@app.route('/onboarding/submit', methods=['POST'])
def onboarding_submit():
    if _is_onboarding_complete():
        flash('Onboarding already completed.', 'info')
        return redirect(url_for('home'))

    form = request.form

    def _clean(key: str) -> str:
        return (form.get(key) or '').strip()

    business_name = _clean('business_name')
    owner_name = _clean('owner_name')
    phone = _clean('phone')
    email = _clean('email')
    address = _clean('address')
    upi_id = _clean('upi_id')
    gstin = _clean('gstin').upper()
    business_type = _clean('business_type')
    pan = _clean('pan').upper()

    bank_account_number = _clean('bank_account_number')
    confirm_bank_account_number = _clean('confirm_bank_account_number')
    ifsc_code = _clean('ifsc_code').upper()
    account_holder_name = _clean('account_holder_name') or business_name
    branch_name = _clean('branch_name')
    bank_name = _clean('bank_name')

    skip_bank = form.get('skip_bank', 'false').lower() == 'true'

    errors: List[str] = []
    if not business_name:
        errors.append('Business name is required.')
    if not owner_name:
        errors.append('Owner name is required.')
    if not phone:
        errors.append('Phone number is required.')

    normalized_account = _normalize_account_number(bank_account_number)
    normalized_confirm = _normalize_account_number(confirm_bank_account_number)

    if not skip_bank:
        if not normalized_account:
            errors.append('Bank account number is required or choose skip.')
        elif normalized_account != normalized_confirm:
            errors.append('Bank account numbers do not match.')

    if errors:
        for err in errors:
            flash(err, 'danger')
        return redirect(url_for('onboarding'))

    now = datetime.now(timezone.utc)
    info_payload = loading_info()
    data_section = _default_info_sections(now)

    # Business details
    data_section['business'].update({
        'name': business_name,
        'owner': owner_name,
        'address': address,
        'email': email,
        'phone': phone,
        'gstin': gstin,
        'upi_id': upi_id,
        'upi_name': owner_name or business_name,
        'businessType': business_type,
        'pan': pan,
    })

    # Bank details (optional)
    if not skip_bank:
        data_section['bank'].update({
            'account_name': account_holder_name,
            'bank_name': bank_name,
            'branch': branch_name,
            'account_number': normalized_account,
            'ifsc': ifsc_code,
            'bhim': phone,
        })

    # Payment / UPI info
    data_section['upi_info'].update({
        'upi_id': upi_id,
        'upi_name': owner_name or business_name,
    })
    payment_methods = [m for m in data_section['payment'].get('methods', [])]
    if upi_id:
        if 'UPI' not in (method.upper() for method in payment_methods):
            payment_methods.insert(0, 'UPI')
        data_section['payment']['upi_qr_enabled'] = True
    else:
        payment_methods = [m for m in payment_methods if m.upper() != 'UPI'] or ['Bank Transfer']
        data_section['payment']['upi_qr_enabled'] = False
    data_section['payment']['methods'] = payment_methods

    # Account defaults / meta
    data_section['account_defaults']['start_date'] = now.strftime('%Y-%m-%dT%H:%M:%SZ')
    data_section['meta']['version'] = APP_VERSION
    data_section['meta']['created_on'] = now.strftime('%Y-%m-%dT%H:%M:%SZ')

    info_payload['data'] = data_section
    info_payload['onboarding_complete'] = True
    info_payload['last_updated'] = now.strftime('%d %B %Y')
    info_payload.setdefault('created_on', now.strftime('%d %B %Y'))
    info_payload.setdefault('app_name', APP_NAME)
    info_payload['version'] = APP_VERSION

    info_path = get_info_json_path()
    try:
        with open(info_path, 'w', encoding='utf-8') as f:
            json.dump(info_payload, f, indent=2, ensure_ascii=False)
    except Exception as exc:
        flash(f'Failed to save onboarding details: {exc}', 'danger')
        return redirect(url_for('onboarding'))

    refresh_info_json()
    flash('Setup complete! You can start generating invoices.', 'success')
    return redirect(url_for('home'))


def get_default_statement_start():
    """Return default statement start date from info.json"""
    tzinfo = tz.gettz(APP_INFO['account_defaults']['timezone'])
    return datetime.strptime(
        APP_INFO['account_defaults']['start_date'], '%Y-%m-%dT%H:%M:%SZ'
    ).replace(tzinfo=tzinfo)


# Call this AFTER importing models, so metadata is populated
with app.app_context():
    _ensure_db_initialized()


# --- routes continue below as usual ---

# Helpers for statement engine
def _parse_date(date_str):
    """Parse 'YYYY-MM-DD' into date or return None."""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        return None


@app.route('/recover')
def recover_page():
    deleted_customers = customer.query.filter_by(isDeleted=True).all()
    deleted_invoices = invoice.query.filter_by(isDeleted=True).all()
    deleted_transactions = (
        accountingTransaction.query
        .options(joinedload(accountingTransaction.customer))
        .filter(accountingTransaction.is_deleted.is_(True))
        .order_by(accountingTransaction.updated_at.desc())
        .all()
    )
    return render_template(
        'recover.html',
        deleted_customers=deleted_customers,
        deleted_invoices=deleted_invoices,
        deleted_transactions=deleted_transactions,
    )


@app.route('/recover_customer/<int:id>')
def recover_customer(id):
    cust = customer.query.get_or_404(id)
    cust.isDeleted = False
    db.session.commit()
    flash('Customer recovered successfully.', 'success')
    return redirect(url_for('recover_page'))


@app.route('/more')
def more():
    supabase_meta = APP_INFO.get('supabase', {})
    last_sync_raw = supabase_meta.get('last_incremental_uploaded') or supabase_meta.get('last_uploaded')
    last_sync_display = _format_sync_timestamp(last_sync_raw)

    latest_backup = _latest_backup_path()
    if latest_backup:
        backup_iso = datetime.fromtimestamp(latest_backup.stat().st_mtime, tz=timezone.utc).isoformat()
        last_backup_display = _format_sync_timestamp(backup_iso)
        last_backup_name = latest_backup.name
    else:
        last_backup_display = "Never"
        last_backup_name = None

    return render_template(
        'more.html',
        has_backup_location=bool((APP_INFO.get('file_location') or '').strip()),
        last_sync_display=last_sync_display,
        last_backup_display=last_backup_display,
        last_backup_name=last_backup_name,
        APP_INFO=APP_INFO,
    )


@app.route('/analytics_event', methods=['GET', 'POST'])
def analytics_event():
    try:
        data = request.get_json(silent=True)

        if not data and request.form:
            data = request.form.to_dict(flat=True)

        if not data and request.data:
            try:
                data = json.loads(request.data.decode('utf-8') or '{}')
            except json.JSONDecodeError:
                data = {}

        normalized = _clean_analytics_payload(data or {})

        if not normalized:
            # Gracefully acknowledge empty analytics pings without treating them as errors
            return jsonify({"status": "ignored", "message": "No analytics payload supplied."}), 204

        # Call the logger in analytics_tracking.py
        log_user_event(normalized)

        return jsonify({"status": "success"}), 200
    except Exception as e:
        print(f"[warn] Analytics log failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/analytics')
def analytics():
    # Get sales trends for day, month, year, and weekday
    day_labels, day_totals = get_sales_trends("day")
    month_labels, month_totals = get_sales_trends("month")
    year_labels, year_totals = get_sales_trends("year")
    # Fetch weekday-level sales trends
    weekday_labels, weekday_totals = get_sales_trends("weekday")
    customer_names, revenues = get_top_customers()
    one_time, repeat = get_customer_retention()
    daywise_labels, daywise_counts, daywise_totals = get_day_wise_billing()

    return render_template(
        'analytics.html',
        day_labels=day_labels,
        day_totals=day_totals,
        month_labels=month_labels,
        month_totals=month_totals,
        year_labels=year_labels,
        year_totals=year_totals,
        weekday_labels=weekday_labels,
        weekday_totals=weekday_totals,
        customer_names=customer_names,
        revenues=revenues,
        one_time=one_time,
        repeat=repeat,
        daywise_labels=daywise_labels,
        daywise_counts=daywise_counts,
        daywise_totals=daywise_totals,
    )


@app.route('/about_user', methods=['GET', 'POST'])
def about_user():
    customer_id = request.args.get('customer_id')
    cust = customer.query.filter_by(id=customer_id, isDeleted=False).first_or_404()
    data = {
        'id': cust.id,
        'name': cust.name,
        'email': cust.email,
        'company': cust.company,
        'phone': cust.phone,
        'gst': cust.gst,
        'address': cust.address,
        'businessType': cust.businessType
    }
    return render_template(
        'about_user.html',
        data=data,
        app_info=APP_INFO

    )


@app.route('/edit_user/<int:customer_id>', methods=['GET', 'POST'])
def edit_user(customer_id):
    cust = customer.query.filter_by(id=customer_id, isDeleted=False).first_or_404()

    if request.method == 'GET':
        return render_template('edit_user.html', customer=cust)

    # POST logic: update values
    name = request.form.get('name', '').strip()
    phone = request.form.get('phone', '').strip()
    address = request.form.get('address', '').strip()
    gst = request.form.get('gst', '').strip()
    email = request.form.get('email', '').strip()
    businessType = request.form.get('businessType', '').strip()
    company = request.form.get('company', '').strip()

    if not name or not phone:
        flash('Name and Phone are required fields.', 'warning')
        return redirect(request.referrer or url_for('about_user', customer_id=customer_id))

    cust.name = name
    cust.phone = phone
    cust.address = address
    cust.gst = gst
    cust.email = email
    cust.businessType = businessType
    cust.company = company

    try:
        db.session.commit()
        flash('Customer updated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating customer: {e}', 'danger')

    next_url = request.form.get('next') or url_for('about_user', customer_id=customer_id)
    return redirect(next_url)


@app.route('/recover_invoice/<int:id>')
def recover_invoice(id):
    inv = invoice.query.get_or_404(id)
    inv.isDeleted = False
    db.session.commit()
    flash('Invoice recovered successfully.', 'success')
    return redirect(url_for('recover_page'))


@app.route('/recover_transaction/<int:txn_id>')
def recover_transaction(txn_id):
    txn = accountingTransaction.query.get_or_404(txn_id)
    txn.is_deleted = False
    db.session.commit()
    flash('Transaction recovered successfully.', 'success')
    return redirect(url_for('recover_page'))


# Custom Jinja filter to format dates as DD Month YYYY (e.g., '14 October 2025')
@app.template_filter('datetimeformat')
def datetimeformat(value, format='%d %B %Y'):
    """Safely format a date string or datetime into a readable form (e.g., '14 October 2025')."""
    if not value:
        return ''
    try:
        if isinstance(value, str):
            # Try parsing both full and simple date formats
            for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S'):
                try:
                    value = datetime.strptime(value, fmt)
                    break
                except ValueError:
                    continue
        return value.strftime(format)
    except Exception:
        return str(value)


@app.route('/_flash_test')
def _flash_test():
    flash('Flash works!', 'success')
    return redirect(url_for('view_customers'))


# Home Route
@app.route('/')
def home():
    session['persistent_notice'] = None
    supabase_meta = APP_INFO.get('supabase', {})
    last_incremental = supabase_meta.get('last_incremental_uploaded')
    last_full = supabase_meta.get('last_uploaded')
    display_last = _format_sync_timestamp(last_incremental or last_full)
    return render_template('home.html', last_uploaded=display_last)


@app.route('/config', methods=['GET', 'POST'])
def config():
    info_path = get_info_json_path()
    layout_config = layoutConfig.get_or_create()
    layout_sizes = layout_config.get_sizes()

    # --- Load existing info.json ---
    with open(info_path, 'r', encoding='utf-8') as f:
        info_data = json.load(f)

    app_info = info_data.get("data", {})
    if not isinstance(app_info, dict):
        app_info = {}
    if 'file_location' not in app_info:
        app_info['file_location'] = ''
    # Merge legacy Cloud Settings block into supabase if present
    legacy_cloud = app_info.pop('Cloud Settings', None)
    if legacy_cloud:
        supabase_section = app_info.setdefault('supabase', {})
        if isinstance(legacy_cloud, dict):
            supabase_section.update(legacy_cloud)

    # --- Handle POST (Save Changes) ---
    if request.method == 'POST':
        section = request.form.get('section')
        if not section:
            flash('No section specified for update.', 'warning')
            return redirect(url_for('config'))

        # Get editable section data
        updates = {}
        for key, val in request.form.items():
            if key not in ('section',):
                updates[key] = val.strip()

        # Apply updates to correct section
        if section == 'file_location':
            new_path = updates.get('file_location') or updates.get('value') or ''
            app_info['file_location'] = new_path.strip()
        elif section == 'invoice_visual':
            business_section = app_info.setdefault('business', {})
            color_mode = (updates.get('logo_color_mode') or '').lower()
            if color_mode:
                business_section['logo_color_mode'] = color_mode

            size_fields = ('header', 'customer', 'invoice_info', 'table', 'totals', 'payment', 'footer')
            sizes_changed = False
            for field in size_fields:
                raw_value = request.form.get(field)
                if raw_value is None:
                    continue
                try:
                    new_value = int(raw_value)
                except (TypeError, ValueError):
                    continue
                if layout_sizes.get(field) != new_value:
                    layout_sizes[field] = new_value
                    sizes_changed = True
            if sizes_changed:
                layout_config.set_sizes(layout_sizes)
                db.session.commit()
        elif section in app_info:
            if isinstance(app_info[section], dict):
                target_section = app_info[section]
                filtered_updates = {}
                for key, new_value in updates.items():
                    existing_value = target_section.get(key)
                    if isinstance(existing_value, str) and new_value == existing_value:
                        continue
                    if new_value == '' and key in target_section:
                        continue
                    filtered_updates[key] = new_value
                if filtered_updates:
                    target_section.update(filtered_updates)
            elif isinstance(app_info[section], list):
                # handle lists (e.g., services textarea)
                lines = updates.get('services', '').splitlines()
                app_info[section] = [ln.strip() for ln in lines if ln.strip()]
            else:
                # simple scalar fields
                value = updates.get(section) or updates.get('value')
                app_info[section] = value.strip() if isinstance(value, str) else updates
        else:
            app_info[section] = updates

        # Update timestamp + save to file
        info_data['data'] = app_info
        info_data['last_updated'] = datetime.now(timezone.utc).strftime("%d %B %Y")

        try:
            with open(info_path, 'w', encoding='utf-8') as f:
                json.dump(info_data, f, indent=2, ensure_ascii=False)
            section_label = section.replace('_', ' ').title()
            flash(f"{section_label} updated successfully!", "success")
            refresh_info_json()
        except Exception as e:
            flash(f"Error saving changes: {e}", "danger")

        # Reload updated version
        return redirect(url_for('config'))

    # --- Default (GET) view ---
    return render_template('config_editor.html', app_info=app_info,
                           last_updated=info_data['last_updated'],
                           created_on=info_data['created_on'],
                           layout_sizes=layout_sizes)


@app.route('/config/refresh', methods=['POST'])
def config_refresh():
    try:
        refresh_info_json()
        flash('Account settings reloaded from info.json.', 'success')
    except Exception as exc:
        flash(f'Unable to refresh settings: {exc}', 'danger')
    return redirect(url_for('config'))


def _accounting_totals(sort_by='balance', sort_dir='desc'):
    base_query = db.session.query
    income_total = base_query(func.coalesce(func.sum(accountingTransaction.amount), 0.0)).filter(
        accountingTransaction.txn_type == 'income',
        accountingTransaction.is_deleted.is_(False)
    ).scalar() or 0.0

    expense_total = base_query(func.coalesce(func.sum(accountingTransaction.amount), 0.0)).filter(
        accountingTransaction.txn_type == 'expense',
        accountingTransaction.is_deleted.is_(False)
    ).scalar() or 0.0

    outstanding_invoice_rows = _outstanding_invoice_rows()
    general_payments = _general_customer_payments()
    customer_expenses = _customer_expenses()
    outstanding_entries = _group_outstanding_by_customer(
        outstanding_invoice_rows,
        general_payments,
        customer_expenses,
        sort_by,
        sort_dir
    )
    outstanding_total = sum(entry['balance'] for entry in outstanding_entries)

    return {
        'income_total': income_total,
        'expense_total': expense_total,
        'net_flow': income_total - expense_total,
        'outstanding_total': outstanding_total,
        'outstanding_entries': outstanding_entries,
        'outstanding_invoices_raw': outstanding_invoice_rows,
        'sort_by': sort_by,
        'sort_dir': sort_dir,
    }


def _outstanding_invoice_rows():
    payments_subq = (
        db.session.query(
            accountingTransaction.invoice_no.label('invoice_no'),
            func.coalesce(func.sum(accountingTransaction.amount), 0.0).label('paid_amount')
        )
        .filter(
            accountingTransaction.is_deleted.is_(False),
            accountingTransaction.txn_type == 'income',
            accountingTransaction.invoice_no.isnot(None)
        )
        .group_by(accountingTransaction.invoice_no)
        .subquery()
    )

    rows = (
        db.session.query(
            invoice.invoiceId.label('invoice_no'),
            invoice.createdAt,
            invoice.totalAmount,
            customer.id.label('customer_id'),
            customer.name.label('customer_name'),
            customer.company.label('customer_company'),
            func.coalesce(payments_subq.c.paid_amount, 0.0).label('paid_amount')
        )
        .join(customer, invoice.customerId == customer.id)
        .outerjoin(payments_subq, payments_subq.c.invoice_no == invoice.invoiceId)
        .filter(
            invoice.isDeleted.is_(False),
            or_(invoice.payment.is_(False), invoice.payment.is_(None))
        )
        .order_by(invoice.createdAt.desc())
    )

    invoice_rows = []
    for row in rows:
        balance = float(max((row.totalAmount or 0) - (row.paid_amount or 0), 0))
        if balance <= 0.01:
            continue
        invoice_rows.append({
            'invoice_no': row.invoice_no,
            'created_at': row.createdAt,
            'customer_id': row.customer_id,
            'customer': row.customer_name or row.customer_company or 'Customer',
            'company': row.customer_company,
            'total': float(row.totalAmount or 0),
            'paid': float(row.paid_amount or 0),
            'balance': balance,
        })
    return invoice_rows


def _general_customer_payments():
    rows = (
        db.session.query(
            accountingTransaction.customerId,
            func.coalesce(func.sum(accountingTransaction.amount), 0.0).label('amt')
        )
        .filter(
            accountingTransaction.is_deleted.is_(False),
            accountingTransaction.txn_type == 'income',
            accountingTransaction.invoice_no.is_(None),
            accountingTransaction.customerId.isnot(None)
        )
        .group_by(accountingTransaction.customerId)
        .all()
    )
    return {row.customerId: float(row.amt or 0) for row in rows}


def _customer_expenses():
    rows = (
        db.session.query(
            accountingTransaction.customerId,
            func.coalesce(func.sum(accountingTransaction.amount), 0.0).label('amt')
        )
        .filter(
            accountingTransaction.is_deleted.is_(False),
            accountingTransaction.txn_type == 'expense',
            accountingTransaction.customerId.isnot(None)
        )
        .group_by(accountingTransaction.customerId)
        .all()
    )
    return {row.customerId: float(row.amt or 0) for row in rows}


def _group_outstanding_by_customer(invoice_rows, general_payments, customer_expenses, sort_by='balance', sort_dir='desc'):
    grouped = {}
    for entry in invoice_rows:
        cust_id = entry.get('customer_id')
        key = cust_id or (entry['customer'], entry.get('company'))
        bucket = grouped.setdefault(key, {
            'customer_id': cust_id,
            'customer': entry['customer'],
            'company': entry.get('company'),
            'total': 0.0,
            'paid': 0.0,
            'expenses': 0.0,
            'invoice_count': 0,
            'latest_invoice_date': entry.get('created_at'),
        })
        bucket['total'] += entry['total']
        bucket['paid'] += entry['paid']
        bucket['invoice_count'] += 1
        existing_date = bucket.get('latest_invoice_date')
        created_at = entry.get('created_at')
        if existing_date is None or (created_at and created_at > existing_date):
            bucket['latest_invoice_date'] = created_at
    existing_ids = {bucket['customer_id'] for bucket in grouped.values() if bucket.get('customer_id')}
    missing_expense_ids = [cid for cid in customer_expenses.keys() if cid and cid not in existing_ids]
    missing_lookup = {}
    if missing_expense_ids:
        customer_rows = customer.query.filter(customer.id.in_(missing_expense_ids)).all()
        missing_lookup = {row.id: row for row in customer_rows}
    for cid in missing_expense_ids:
        cust_obj = missing_lookup.get(cid)
        grouped[cid] = {
            'customer_id': cid,
            'customer': cust_obj.name if cust_obj else 'Customer',
            'company': cust_obj.company if cust_obj else None,
            'total': 0.0,
            'paid': 0.0,
            'expenses': 0.0,
            'invoice_count': 0,
            'latest_invoice_date': None,
        }

    for bucket in grouped.values():
        cust_id = bucket.get('customer_id')
        general = general_payments.get(cust_id)
        if general:
            bucket['paid'] += general

        expense_sum = customer_expenses.get(cust_id)
        if expense_sum:
            bucket['expenses'] += expense_sum

        total_due = bucket['total'] + bucket['expenses']
        bucket['balance'] = max(total_due - bucket['paid'], 0.0)

    result = list(grouped.values())
    sort_key_map = {
        'invoices': lambda r: r['invoice_count'],
        'balance': lambda r: r['balance'],
        'expenses': lambda r: r['expenses'],
        'paid': lambda r: r['paid'],
        'invoiced': lambda r: r['total'],
    }
    sort_field = sort_key_map.get(sort_by, sort_key_map['balance'])
    reverse = (sort_dir.lower() != 'asc')
    result.sort(key=sort_field, reverse=reverse)
    return result


def _ensure_business_expense_customer():
    cust = (
        customer.query
        .filter(
            customer.name == "Business Expense",
            customer.isDeleted == False
        )
        .first()
    )
    if cust:
        return cust
    synthetic_phone = f"EXP-{int(datetime.now(timezone.utc).timestamp())}"
    cust = customer(
        name="Business Expense",
        company="Internal Ledger",
        phone=synthetic_phone,
        email="",
        gst="",
        address="",
        businessType="Expense",
        isDeleted=False
    )
    db.session.add(cust)
    db.session.flush()
    return cust


def _persist_accounting_transaction(form, existing_txn=None):
    previous_invoice_no = existing_txn.invoice_no if existing_txn else None
    previous_txn_type = existing_txn.txn_type if existing_txn else None
    txn_type = (form.get('txn_type') or 'income').strip().lower()
    if txn_type not in ('income', 'expense'):
        txn_type = 'income'

    amount_raw = (form.get('amount') or '').strip()
    try:
        amount_decimal = Decimal(amount_raw)
    except (InvalidOperation, TypeError):
        return "Enter a valid amount."
    if amount_decimal <= 0:
        return "Amount must be greater than zero."

    customer_id_val = form.get('customer_id')
    customer_obj = None
    if customer_id_val:
        try:
            customer_obj = customer.query.get(int(customer_id_val))
        except (TypeError, ValueError):
            customer_obj = None
        if not customer_obj or customer_obj.isDeleted:
            return "Selected customer could not be found."

    customer_name = None
    if txn_type == 'income' and not customer_obj:
        return "Select a customer for payments received."
    if txn_type == 'expense' and not customer_obj:
        customer_obj = _ensure_business_expense_customer()

    invoice_no = (form.get('invoice_no') or '').strip() or None
    invoice_obj = None
    if invoice_no:
        invoice_obj = invoice.query.filter_by(invoiceId=invoice_no).first()
        if not invoice_obj:
            return "Invoice number could not be located."
        if customer_obj and invoice_obj.customerId != customer_obj.id:
            return "Invoice does not belong to the selected customer."

    txn_created_at = None
    txn_date_raw = (form.get('txn_date') or '').strip()
    if txn_date_raw:
        try:
            tz_name = (APP_INFO.get('account_defaults') or {}).get('timezone') or DEFAULT_TIMEZONE
            local_tz = tz.gettz(tz_name) or timezone.utc
            parsed_date = datetime.strptime(txn_date_raw, '%Y-%m-%d')
            local_dt = datetime(parsed_date.year, parsed_date.month, parsed_date.day, 12, 0, tzinfo=local_tz)
            txn_created_at = local_dt.astimezone(timezone.utc)
        except Exception:
            txn_created_at = None

    txn_kwargs = {}
    if txn_created_at:
        txn_kwargs['created_at'] = txn_created_at
        txn_kwargs['updated_at'] = txn_created_at

    txn = existing_txn
    if txn:
        txn.customerId = customer_obj.id if customer_obj else None
        txn.amount = float(amount_decimal)
        txn.txn_type = txn_type
        txn.mode = (form.get('mode') or '').strip().lower() or 'cash'
        txn.account = (form.get('account') or '').strip().lower() or 'cash'
        txn.invoice_no = invoice_no
        txn.remarks = (form.get('remarks') or '').strip() or None
        if txn_created_at:
            txn.created_at = txn_created_at
        txn.updated_at = datetime.now(timezone.utc)
        db.session.flush()
    else:
        txn = accountingTransaction(
            customerId=customer_obj.id if customer_obj else None,
            amount=float(amount_decimal),
            txn_type=txn_type,
            mode=(form.get('mode') or '').strip().lower() or 'cash',
            account=(form.get('account') or '').strip().lower() or 'cash',
            invoice_no=invoice_no,
            remarks=(form.get('remarks') or '').strip() or None,
            **txn_kwargs
        )
        db.session.add(txn)
        db.session.flush()

    if existing_txn:
        for entry in list(existing_txn.expense_items):
            db.session.delete(entry)

    if txn_type == 'expense':
        descriptions = form.getlist('expense_desc[]')
        amounts = form.getlist('expense_amount[]')
        for idx, desc in enumerate(descriptions):
            desc_text = (desc or '').strip()
            if not desc_text:
                continue
            amount_val = None
            if idx < len(amounts):
                amt_raw = (amounts[idx] or '').strip()
                if amt_raw:
                    try:
                        amount_val = float(Decimal(amt_raw))
                    except (InvalidOperation, TypeError):
                        amount_val = None
            expense_entry = expenseItem(
                transactionId=txn.id,
                description=desc_text,
                amount=amount_val
            )
            db.session.add(expense_entry)

    invoices_to_sync = set()
    if txn_type == 'income' and invoice_no:
        invoices_to_sync.add(invoice_no)
    if previous_txn_type == 'income' and previous_invoice_no:
        invoices_to_sync.add(previous_invoice_no)

    for inv_code in invoices_to_sync:
        _sync_invoice_payment_flag(inv_code)

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return f"Unable to record transaction: {exc}"
    return None


def _sync_invoice_payment_flag(invoice_code: Optional[str]):
    if not invoice_code:
        return
    target_invoice = invoice.query.filter_by(invoiceId=invoice_code, isDeleted=False).first()
    if not target_invoice:
        return
    paid_amount = (
        db.session.query(func.coalesce(func.sum(accountingTransaction.amount), 0.0))
        .filter(
            accountingTransaction.invoice_no == invoice_code,
            accountingTransaction.txn_type == 'income',
            accountingTransaction.is_deleted.is_(False)
        )
        .scalar()
        or 0.0
    )
    invoice_total = float(target_invoice.totalAmount or 0.0)
    target_invoice.payment = paid_amount >= max(invoice_total, 0.0) - 0.01


# Accounting dashboard
@app.route('/accounting', methods=['GET', 'POST'])
def accounting_dashboard():
    sort_by = (request.args.get('sort') or 'balance').lower()
    sort_dir = (request.args.get('dir') or 'desc').lower()
    if sort_by not in {'balance', 'expenses', 'paid', 'invoices', 'invoiced'}:
        sort_by = 'balance'
    if sort_dir not in {'asc', 'desc'}:
        sort_dir = 'desc'

    if request.method == 'POST':
        next_url = request.form.get('next_url')
        allowed_next = {url_for('accounting_dashboard'), url_for('accounting_transactions_list')}
        if next_url not in allowed_next:
            next_url = url_for('accounting_dashboard')
        error = _persist_accounting_transaction(request.form)
        if error:
            db.session.rollback()
            flash(error, 'danger')
        else:
            flash('Transaction recorded successfully.', 'success')
        return redirect(next_url)

    totals = _accounting_totals(sort_by=sort_by, sort_dir=sort_dir)
    outstanding = totals['outstanding_entries']
    customers_list = customer.alive().order_by(customer.name.asc()).all()
    recent_transactions = (
        accountingTransaction.query
        .options(joinedload(accountingTransaction.expense_items), joinedload(accountingTransaction.customer))
        .filter(accountingTransaction.is_deleted.is_(False))
        .order_by(accountingTransaction.id.desc())
        .limit(6)
        .all()
    )

    invoice_choices = []
    seen_invoices = set()
    for entry in totals.get('outstanding_invoices_raw', []):
        inv_code = entry.get('invoice_no')
        if inv_code and inv_code not in seen_invoices:
            seen_invoices.add(inv_code)
            invoice_choices.append(inv_code)

    payment_modes = ['cash', 'bank', 'upi']
    account_options = ['cash', 'savings', 'current']
    business_expense_id = _ensure_business_expense_customer().id

    return render_template(
        'accounting.html',
        totals=totals,
        outstanding=outstanding,
        customers=customers_list,
        recent_transactions=recent_transactions,
        invoice_choices=invoice_choices,
        payment_modes=payment_modes,
        account_options=account_options,
        current_sort=sort_by,
        current_dir=sort_dir,
        business_expense_id=business_expense_id,
    )


@app.route('/accounting/quick_clear/<int:customer_id>', methods=['POST'])
def accounting_quick_clear(customer_id):
    cust = customer.query.filter_by(id=customer_id, isDeleted=False).first()
    if not cust:
        flash('Customer not found.', 'warning')
        return redirect(url_for('accounting_dashboard'))

    snapshot = _customer_financial_snapshot(customer_id)
    balance = float(snapshot.get('balance') or 0.0)
    if balance <= 0:
        flash('No outstanding balance to clear for this customer.', 'info')
        return redirect(url_for('accounting_dashboard'))

    txn = accountingTransaction(
        customerId=customer_id,
        amount=balance,
        txn_type='income',
        mode='bank',
        account='current',
        remarks='Quick clear via dashboard',
    )
    db.session.add(txn)
    cleared_invoices = (
        invoice.query
        .filter(invoice.customerId == customer_id, invoice.isDeleted.is_(False))
        .all()
    )
    for inv in cleared_invoices:
        inv.payment = True
    try:
        db.session.commit()
        flash(f"Cleared INR: {balance:,.2f} for {cust.name}.", 'success')
    except Exception as exc:
        db.session.rollback()
        flash(f"Unable to clear dues: {exc}", 'danger')

    return redirect(url_for('accounting_dashboard'))


@app.route('/accounting/transactions')
def accounting_transactions_list():
    sort_by = (request.args.get('sort') or 'id').lower()
    sort_dir = (request.args.get('dir') or 'desc').lower()
    if sort_by not in {'date', 'amount', 'customer', 'type', 'id'}:
        sort_by = 'id'
    if sort_dir not in {'asc', 'desc'}:
        sort_dir = 'desc'

    customer_query = (request.args.get('customer') or '').strip()
    date_query = (request.args.get('date') or '').strip()
    amount_query = (request.args.get('amount') or '').strip()

    q = (
        accountingTransaction.query
        .outerjoin(customer)
        .options(joinedload(accountingTransaction.customer))
        .filter(accountingTransaction.is_deleted.is_(False))
    )

    if customer_query:
        like_value = f"%{customer_query.lower()}%"
        q = q.filter(
            or_(
                func.lower(customer.name).like(like_value),
                func.lower(customer.company).like(like_value),
                func.lower(accountingTransaction.txn_id).like(like_value)
            )
        )

    if date_query:
        try:
            parsed_date = datetime.strptime(date_query, '%Y-%m-%d')
            start = parsed_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end = start + timedelta(days=1)
            q = q.filter(accountingTransaction.created_at >= start, accountingTransaction.created_at < end)
        except Exception:
            pass

    if amount_query:
        try:
            amt_value = float(Decimal(amount_query))
            q = q.filter(accountingTransaction.amount == amt_value)
        except (InvalidOperation, ValueError):
            pass

    if sort_by == 'amount':
        sort_column = accountingTransaction.amount
    elif sort_by == 'customer':
        sort_column = func.lower(customer.name)
    elif sort_by == 'type':
        sort_column = accountingTransaction.txn_type
    elif sort_by == 'id':
        sort_column = accountingTransaction.id
    else:
        sort_column = accountingTransaction.created_at

    if sort_dir == 'asc':
        q = q.order_by(sort_column.asc(), accountingTransaction.id.asc())
    else:
        q = q.order_by(sort_column.desc(), accountingTransaction.id.desc())

    transactions = q.all()

    customers_list = customer.alive().order_by(customer.name.asc()).all()
    invoice_choices = []
    seen = set()
    for row in _outstanding_invoice_rows():
        inv = row.get('invoice_no')
        if inv and inv not in seen:
            seen.add(inv)
            invoice_choices.append(inv)
    payment_modes = ['cash', 'bank', 'upi']
    account_options = ['cash', 'savings', 'current']
    business_expense_id = _ensure_business_expense_customer().id

    return render_template(
        'accounting_transactions.html',
        transactions=transactions,
        filter_customer=customer_query,
        filter_date=date_query,
        filter_amount=amount_query,
        current_sort=sort_by,
        current_dir=sort_dir,
        customers=customers_list,
        invoice_choices=invoice_choices,
        payment_modes=payment_modes,
        account_options=account_options,
        business_expense_id=business_expense_id,
    )


@app.route('/accounting/transactions/<int:txn_id>', methods=['GET', 'POST'])
def accounting_transaction_detail(txn_id):
    txn = (
        accountingTransaction.query
        .options(joinedload(accountingTransaction.customer), joinedload(accountingTransaction.expense_items))
        .get_or_404(txn_id)
    )

    if request.method == 'POST':
        if txn.is_deleted:
            flash('Transaction already archived.', 'warning')
        else:
            txn.is_deleted = True
            invoice_code = txn.invoice_no if txn.txn_type == 'income' else None
            _sync_invoice_payment_flag(invoice_code)
            db.session.commit()
            flash('Transaction deleted successfully.', 'success')
        return redirect(url_for('accounting_transactions_list'))

    return render_template('accounting_transaction_detail.html', txn=txn)


@app.route('/accounting/transactions/<int:txn_id>/edit', methods=['GET', 'POST'])
def accounting_transaction_edit(txn_id):
    txn = (
        accountingTransaction.query
        .options(joinedload(accountingTransaction.customer), joinedload(accountingTransaction.expense_items))
        .get_or_404(txn_id)
    )
    if txn.is_deleted:
        flash('Cannot edit an archived transaction.', 'warning')
        return redirect(url_for('accounting_transaction_detail', txn_id=txn_id))

    customers_list = customer.alive().order_by(customer.name.asc()).all()
    invoice_choices = []
    seen = set()
    for row in _outstanding_invoice_rows():
        inv = row.get('invoice_no')
        if inv and inv not in seen:
            seen.add(inv)
            invoice_choices.append(inv)
    if txn.invoice_no and txn.invoice_no not in invoice_choices:
        invoice_choices.append(txn.invoice_no)
    payment_modes = ['cash', 'bank', 'upi']
    account_options = ['cash', 'savings', 'current']
    business_expense_id = _ensure_business_expense_customer().id

    form_data = None
    expense_rows = []

    def _build_expense_rows(source_form):
        if not source_form:
            rows = [
                {
                    'desc': (item.description or ''),
                    'amount': '' if item.amount is None else f"{item.amount:.2f}"
                }
                for item in txn.expense_items
            ]
        else:
            descs = source_form.getlist('expense_desc[]')
            amts = source_form.getlist('expense_amount[]')
            rows = []
            for idx in range(max(len(descs), len(amts))):
                desc_val = descs[idx] if idx < len(descs) else ''
                amt_val = amts[idx] if idx < len(amts) else ''
                rows.append({'desc': desc_val, 'amount': amt_val})
        if not rows:
            rows = [{'desc': '', 'amount': ''}]
        return rows

    if request.method == 'POST':
        form_data = request.form
        error = _persist_accounting_transaction(request.form, existing_txn=txn)
        if error:
            db.session.rollback()
            flash(error, 'danger')
            expense_rows = _build_expense_rows(form_data)
        else:
            flash('Transaction updated successfully.', 'success')
            return redirect(url_for('accounting_transaction_detail', txn_id=txn.id))
    else:
        expense_rows = _build_expense_rows(None)

    return render_template(
        'accounting_transaction_edit.html',
        txn=txn,
        customers=customers_list,
        invoice_choices=invoice_choices,
        payment_modes=payment_modes,
        account_options=account_options,
        business_expense_id=business_expense_id,
        form_data=form_data,
        expense_rows=expense_rows
    )


def _customer_financial_snapshot(customer_id: int) -> dict:
    total_invoiced, invoice_count = (
        db.session.query(
            func.coalesce(func.sum(invoice.totalAmount), 0.0),
            func.count(invoice.id)
        )
        .filter(
            invoice.customerId == customer_id,
            invoice.isDeleted.is_(False)
        )
        .one()
    )

    total_payments = (
        db.session.query(func.coalesce(func.sum(accountingTransaction.amount), 0.0))
        .filter(
            accountingTransaction.customerId == customer_id,
            accountingTransaction.txn_type == 'income',
            accountingTransaction.is_deleted.is_(False)
        )
        .scalar()
        or 0.0
    )

    total_expenses = (
        db.session.query(func.coalesce(func.sum(accountingTransaction.amount), 0.0))
        .filter(
            accountingTransaction.customerId == customer_id,
            accountingTransaction.txn_type == 'expense',
            accountingTransaction.is_deleted.is_(False)
        )
        .scalar()
        or 0.0
    )

    balance = (total_invoiced + total_expenses) - total_payments

    return {
        'customer_id': customer_id,
        'total_invoiced': float(total_invoiced or 0.0),
        'invoice_count': int(invoice_count or 0),
        'total_payments': float(total_payments or 0.0),
        'total_expenses': float(total_expenses or 0.0),
        'balance': float(balance),
    }


def _build_accounting_statement_context(
    start_dt: datetime,
    end_dt: datetime,
    start_date_display,
    end_date_display,
    txn_type_filter: str = 'all',
    customer_query: str = ''
) -> dict:
    """
    Aggregate accounting transactions for the printable accounting statement view.
    Returns totals, breakdowns, and the matching transaction rows.
    """
    txn_filter = txn_type_filter if txn_type_filter in {'income', 'expense'} else 'all'
    customer_filter = (customer_query or '').strip()

    tz_name = (APP_INFO.get('account_defaults') or {}).get('timezone') or DEFAULT_TIMEZONE
    display_tz = tz.gettz(tz_name) or timezone.utc

    q = (
        accountingTransaction.query
        .options(joinedload(accountingTransaction.customer))
        .filter(
            accountingTransaction.is_deleted.is_(False),
            accountingTransaction.created_at >= start_dt,
            accountingTransaction.created_at <= end_dt,
        )
    )

    if txn_filter in {'income', 'expense'}:
        q = q.filter(accountingTransaction.txn_type == txn_filter)

    if customer_filter:
        like_value = f"%{customer_filter.lower()}%"
        q = q.join(customer, accountingTransaction.customerId == customer.id)
        q = q.filter(
            or_(
                func.lower(func.coalesce(customer.name, '')).like(like_value),
                func.lower(func.coalesce(customer.company, '')).like(like_value),
                func.lower(func.coalesce(customer.phone, '')).like(like_value),
            )
        )
    else:
        q = q.outerjoin(customer, accountingTransaction.customerId == customer.id)

    transactions = q.order_by(accountingTransaction.created_at.desc()).all()

    selected_customer = None
    normalized_filter = customer_filter.lower()
    if customer_filter:
        selected_customer = (
            customer.query
            .filter(
                customer.isDeleted == False,
                func.lower(customer.phone) == normalized_filter
            )
            .first()
        )
        if not selected_customer:
            try:
                customer_id_match = int(customer_filter)
            except (TypeError, ValueError):
                customer_id_match = None
            if customer_id_match:
                selected_customer = (
                    customer.query
                    .filter(
                        customer.isDeleted == False,
                        customer.id == customer_id_match
                    )
                    .first()
                )
        if not selected_customer:
            like_value = f"%{normalized_filter}%"
            selected_customer = (
                customer.query
                .filter(
                    customer.isDeleted == False,
                    or_(
                        func.lower(customer.name).like(like_value),
                        func.lower(customer.company).like(like_value)
                    )
                )
                .order_by(customer.name.asc())
                .first()
            )
    if customer_filter and not selected_customer:
        for txn in transactions:
            if txn.customer:
                selected_customer = txn.customer
                break

    income_total = 0.0
    expense_total = 0.0
    income_count = 0
    expense_count = 0
    daily_totals = defaultdict(lambda: {'income': 0.0, 'expense': 0.0})
    mode_totals = defaultdict(lambda: {'income': 0.0, 'expense': 0.0})
    account_totals = defaultdict(lambda: {'income': 0.0, 'expense': 0.0})
    customer_totals = {}

    for txn in transactions:
        amount = float(txn.amount or 0.0)
        if amount <= 0:
            continue
        created_at = txn.created_at or start_dt
        if created_at.tzinfo is None:
            created_at = created_at.replace(tzinfo=timezone.utc)
        local_created = created_at.astimezone(display_tz)
        date_key = local_created.date()
        day_bucket = daily_totals[date_key]

        if txn.txn_type == 'income':
            income_total += amount
            income_count += 1
            day_bucket['income'] += amount
        else:
            expense_total += amount
            expense_count += 1
            day_bucket['expense'] += amount

        mode_label = (txn.mode or 'Unspecified').strip() or 'Unspecified'
        account_label = (txn.account or 'Unspecified').strip() or 'Unspecified'
        mode_bucket = mode_totals[mode_label.upper()]
        account_bucket = account_totals[account_label.title()]
        if txn.txn_type == 'income':
            mode_bucket['income'] += amount
            account_bucket['income'] += amount
        else:
            mode_bucket['expense'] += amount
            account_bucket['expense'] += amount

        cust_key = txn.customerId if txn.customerId else 'unassigned'
        customer_bucket = customer_totals.get(cust_key)
        if not customer_bucket:
            label = txn.customer.name if txn.customer else 'Unassigned'
            customer_bucket = {
                'customer_id': txn.customer.id if txn.customer else None,
                'customer': label,
                'company': txn.customer.company if txn.customer else '',
                'phone': txn.customer.phone if txn.customer else '',
                'income': 0.0,
                'expense': 0.0,
                'transactions': 0,
                'last_txn_at': local_created,
            }
            customer_totals[cust_key] = customer_bucket

        customer_bucket['transactions'] += 1
        if txn.txn_type == 'income':
            customer_bucket['income'] += amount
        else:
            customer_bucket['expense'] += amount

        if local_created > customer_bucket['last_txn_at']:
            customer_bucket['last_txn_at'] = local_created

    def _format_breakdown(source):
        rows = []
        for label, values in source.items():
            total = values['income'] + values['expense']
            rows.append({
                'label': label,
                'income': values['income'],
                'expense': values['expense'],
                'net': values['income'] - values['expense'],
                'total': total,
            })
        rows.sort(key=lambda row: row['total'], reverse=True)
        return rows

    customer_summary = []
    for entry in customer_totals.values():
        entry['net'] = entry['income'] - entry['expense']
        customer_summary.append(entry)
    customer_summary.sort(
        key=lambda row: (row['customer_id'] is None, -row['net'])
    )

    daily_summary = []
    for day, values in sorted(daily_totals.items()):
        daily_summary.append({
            'date': day,
            'income': values['income'],
            'expense': values['expense'],
            'net': values['income'] - values['expense'],
        })

    show_income_columns = income_total > 0.0
    show_expense_columns = expense_total > 0.0

    customer_invoices = []
    customer_payments = []
    customer_adjustments = []
    customer_statement_summary = None
    selected_customer_info = None
    if selected_customer:
        selected_customer_info = {
            'id': selected_customer.id,
            'name': selected_customer.name,
            'company': selected_customer.company,
            'phone': selected_customer.phone,
        }
        invoice_rows = (
            invoice.query
            .filter(
                invoice.customerId == selected_customer.id,
                invoice.isDeleted == False,
                invoice.createdAt >= start_dt,
                invoice.createdAt <= end_dt
            )
            .order_by(invoice.createdAt.desc())
            .all()
        )
        for inv in invoice_rows:
            inv_created = inv.createdAt or start_dt
            if inv_created.tzinfo is None:
                inv_created = inv_created.replace(tzinfo=timezone.utc)
            local_inv = inv_created.astimezone(display_tz)
            customer_invoices.append({
                'invoice_no': inv.invoiceId,
                'date': local_inv,
                'total': float(inv.totalAmount or 0),
            })

        for txn in transactions:
            if txn.customerId != selected_customer.id:
                continue
            local_created = (txn.created_at or start_dt)
            if local_created.tzinfo is None:
                local_created = local_created.replace(tzinfo=timezone.utc)
            local_created = local_created.astimezone(display_tz)
            entry = {
                'txn_id': txn.txn_id,
                'date': local_created,
                'mode': txn.mode or '—',
                'account': txn.account or '—',
                'amount': float(txn.amount or 0),
                'remarks': txn.remarks,
            }
            if txn.txn_type == 'income':
                customer_payments.append(entry)
            else:
                customer_adjustments.append(entry)

        invoice_total = sum(inv['total'] for inv in customer_invoices)
        payment_total = sum(p['amount'] for p in customer_payments)
        adjustment_total = sum(adj['amount'] for adj in customer_adjustments)
        customer_statement_summary = {
            'invoice_total': round(invoice_total, 2),
            'payment_total': round(payment_total, 2),
            'adjustment_total': round(adjustment_total, 2),
            'balance': round(invoice_total + adjustment_total - payment_total, 2),
            'invoice_count': len(customer_invoices),
            'payment_count': len(customer_payments),
        }

    context = {
        'transactions': transactions,
        'income_total': round(income_total, 2),
        'expense_total': round(expense_total, 2),
        'net_flow': round(income_total - expense_total, 2),
        'transaction_count': len(transactions),
        'income_count': income_count,
        'expense_count': expense_count,
        'filters': {
            'start': start_date_display.isoformat(),
            'end': end_date_display.isoformat(),
            'customer': customer_filter,
            'type': txn_filter,
        },
        'customer_filter': customer_filter,
        'selected_type': txn_filter,
        'customer_summary': customer_summary,
        'mode_breakdown': _format_breakdown(mode_totals),
        'account_breakdown': _format_breakdown(account_totals),
        'daily_summary': daily_summary,
        'display_timezone': tz_name,
        'generated_at': datetime.now(timezone.utc).astimezone(display_tz),
        'start_date': start_date_display,
        'end_date': end_date_display,
        'start_local': start_dt.astimezone(display_tz),
        'end_local': end_dt.astimezone(display_tz),
        'overall_totals': _accounting_totals(),
        'customer_invoices': customer_invoices,
        'selected_customer': selected_customer_info,
        'customer_payments': customer_payments,
        'customer_adjustments': customer_adjustments,
        'customer_statement_summary': customer_statement_summary,
        'show_income_columns': show_income_columns,
        'show_expense_columns': show_expense_columns,
    }
    context['statement_range_label'] = (
        f"{context['start_local'].strftime('%d %b %Y')} — {context['end_local'].strftime('%d %b %Y')}"
    )
    return context


@app.route('/accounting/customer_summary/<int:customer_id>')
def accounting_customer_summary(customer_id):
    cust = customer.query.filter_by(id=customer_id, isDeleted=False).first()
    if not cust:
        return jsonify({'error': 'Customer not found'}), 404
    summary = _customer_financial_snapshot(customer_id)
    summary['customer_name'] = cust.name
    summary['company'] = cust.company
    return jsonify(summary)


@app.route('/accounting/amount_to_words')
def accounting_amount_to_words():
    amount = request.args.get('amount', '0')
    return jsonify({'words': amount_to_words(amount)})


# customers page (temperory placeholder)
@app.route('/create_customers', methods=['GET', 'POST'])
def add_customers():
    if request.method == 'POST':
        use_auto = bool(request.form.get('use_auto_id'))
        phone = (request.form.get('phone') or '').strip()
        name = (request.form.get('name') or '').strip()
        company = (request.form.get('company') or '').strip()
        email = (request.form.get('email') or '').strip()
        gst = (request.form.get('gst') or '').strip()
        address = (request.form.get('address') or '').strip()
        businessType = (request.form.get('businessType') or '').strip()

        # --- Basic validation ---
        if not name or (not use_auto and not phone):
            return render_template(
                'add_customer.html',
                success=False,
                duplicate=False,
                error='Name is required. Phone is required unless you use computer-generated ID.',
                # sticky values
                name=name, company=company, phone=phone, email=email,
                gst=gst, address=address, businessType=businessType,
                use_auto_id=use_auto
            )

        # --- Duplicate checks (exclude soft-deleted if you have that flag) ---
        not_deleted = getattr(customer, 'isDeleted', False) == False

        # 1) Phone duplicate (only when not using auto-id)
        if not use_auto and phone:
            existing_phone = (customer.query
                              .filter(func.lower(customer.phone) == phone.lower(), not_deleted)
                              .first())
            if existing_phone:
                return render_template(
                    'add_customer.html',
                    duplicate=True,
                    error='A customer with this phone already exists.',
                    name=name, company=company, phone=phone, email=email,
                    gst=gst, address=address, businessType=businessType,
                    use_auto_id=use_auto
                )

        # 2) Company+Name duplicate (case-insensitive)
        if company and name:
            existing_pair = (customer.query
                             .filter(func.lower(customer.company) == company.lower(),
                                     func.lower(customer.name) == name.lower(),
                                     not_deleted)
                             .first())
            if existing_pair:
                return render_template(
                    'add_customer.html',
                    duplicate=True,
                    error='A customer with the same Company + Name already exists.',
                    name=name, company=company, phone=phone, email=email,
                    gst=gst, address=address, businessType=businessType,
                    use_auto_id=use_auto
                )

        # --- Create customer ---
        if not use_auto and phone:
            # Real phone path
            c = customer(
                name=name, company=company, phone=phone, email=email,
                gst=gst, address=address, businessType=businessType
            )
            db.session.add(c)
            db.session.commit()
            # add alert
            flash('New Customer Created successfully.', 'success')
            return redirect(url_for('about_user', customer_id=c.id))

        # Auto-ID path (no real phone or toggle checked)
        temp_phone = f"ID-TEMP-{uuid.uuid4().hex[:8]}"
        c = customer(
            name=name, company=company, phone=temp_phone, email=email,
            gst=gst, address=address, businessType=businessType
        )

        db.session.add(c)
        db.session.flush()  # get c.id
        c.phone = _format_customer_id(c.id)  # e.g., ID-000123
        db.session.commit()
        # add alert
        flash('New Customer Created successfully.', 'success')

        return redirect(url_for('about_user', customer_id=c.id))

    # GET -> render blank form
    return render_template('add_customer.html')


@app.route('/delete_customer/<int:cid>', methods=['GET', 'POST'])
def delete_customer(cid):
    # Load customer
    c = db.session.get(customer, cid)
    if not c:
        flash("Customer not found", 'warning')
        return redirect(url_for('view_customers'))

    # If already deleted, bail gracefully
    if hasattr(c, 'isDeleted') and c.isDeleted:
        flash('Customer already deleted.', 'info')
        return redirect(url_for('view_customers'))

    # Live invoices (exclude soft-deleted)
    inv_q = invoice.query.filter(
        invoice.customerId == cid,
        getattr(invoice, 'isDeleted', False) == False
    )
    inv_count = inv_q.count()
    total_billed = db.session.query(
        func.coalesce(func.sum(invoice.totalAmount), 0.0)
    ).filter(
        invoice.customerId == cid,
        getattr(invoice, 'isDeleted', False) == False
    ).scalar() or 0.0

    # If POST with confirm flag -> cascade soft-delete customer + invoices
    if request.method == 'POST' and request.form.get('confirm') == '1':
        if hasattr(c, 'isDeleted'):
            c.isDeleted = True
        # Soft delete all invoices for this customer
        invs = invoice.query.filter_by(customerId=cid).all()
        for inv in invs:
            if hasattr(inv, 'isDeleted'):
                inv.isDeleted = True
                inv.deletedAt = datetime.now(timezone.utc)
        db.session.commit()
        # add alert,
        flash('Customer and related invoices deleted successfully.', 'danger')
        return redirect(url_for('view_customers'))

    # GET: If invoices exist but not confirmed yet -> show confirm page
    if request.method == 'GET' and inv_count > 0:
        return render_template(
            'confirm_delete_customer.html',
            customer=c,
            inv_count=inv_count,
            total_billed=total_billed
        )

    # No invoices -> delete immediately (GET or POST)
    if hasattr(c, 'isDeleted'):
        c.isDeleted = True
        db.session.commit()
        flash('Customer deleted.', 'danger')
    else:
        flash('Delete not available in this build.', 'warning')

    return redirect(url_for('view_customers'))


@app.route('/add_inventory', methods=['GET', 'POST'])
def add_inventory():
    # this function will be used to create a new inventory item.
    if request.method == 'POST':
        # Read and normalize inputs
        name = (request.form.get('name') or '').strip()
        unit_price_raw = (request.form.get('unitPrice') or '').strip()
        qty_raw = (request.form.get('quantity') or '').strip()
        tax_raw = (request.form.get('taxPercentage') or '').strip()

        # Basic validation
        if not name:
            return render_template('add_inventory.html', success=False, error='Item name is required.')

        try:
            unit_price = float(unit_price_raw or 0)
        except ValueError:
            return render_template('add_inventory.html', success=False, error='Unit price must be a number.')

        try:
            qty = int(qty_raw or 10000)
        except ValueError:
            return render_template('add_inventory.html', success=False, error='Quantity must be an integer.')

        try:
            tax_pct = float(tax_raw or 0)
        except ValueError:
            return render_template('add_inventory.html', success=False, error='Tax % must be a number.')

        # Duplicate check by name (case-insensitive)
        existing = item.query.filter(func.lower(item.name) == name.lower()).first()
        if existing:
            return render_template('add_inventory.html', duplicate=True)

        # Create item; SKU auto-assigned by model's before_insert listener
        new_item = item(
            name=name,
            unitPrice=unit_price,
            quantity=qty,
            taxPercentage=tax_pct
        )
        db.session.add(new_item)
        db.session.commit()
        # add alert
        flash('Item added successfully.', 'success')
        return render_template('add_inventory.html', success=True)

    return render_template('add_inventory.html')


@app.route('/select_customer', methods=['GET', 'POST'])
def select_customer():
    if request.method == 'POST':
        # User clicked Select on a customer row; the form sends phone
        phone = request.form.get('customer')
        sel = (customer.query
               .filter(customer.isDeleted == False, customer.phone == phone)
               .first_or_404())
        return _render_create_bill(customer=sel, inventory=item.query.all())

    # GET: either search or show recent
    q = (request.args.get('q') or '').strip()
    base = customer.query.filter(customer.isDeleted == False)
    if q:
        like = f"%{q}%"
        customers = (base.filter((customer.phone.ilike(like)) |
                                 (customer.name.ilike(like)) |
                                 (customer.company.ilike(like)))
                     .order_by(customer.id.desc())
                     .limit(100)
                     .all())
    else:
        customers = (base.order_by(customer.id.desc())
                     .limit(25)
                     .all())

    return render_template('select_customer.html', customers=customers)


@app.route('/view_inventory')
def view_inventory():
    query = (request.args.get('q') or '').lower()
    inventory = item.query.all()

    if query:
        inventory = [
            it for it in inventory
            if query in it.name.lower() or (it.sku is not None and query in str(it.sku).lower())
        ]

    return render_template('view_inventory.html', inventory=inventory)


@app.route('/api/statements', methods=['GET'])
def api_statements_summary():
    """JSON summary for dashboards.
    Query params same as /statements (scope/year/month/start/end/phone).
    Returns: {range, totals, per_customer, per_day, per_month}
    """
    # Delegate date-range parsing to the /statements logic by reusing code
    scope = (request.args.get('scope') or 'custom').lower()
    phone = request.args.get('phone')
    today = datetime.now().date()
    if scope == 'year':
        year = int(request.args.get('year') or today.year)
        start_date = datetime(year, 1, 1).date()
        end_date = datetime(year, 12, 31).date()
    elif scope == 'month':
        year = int(request.args.get('year') or today.year)
        month = int(request.args.get('month') or today.month)
        start_date = datetime(year, month, 1).date()
        end_date = datetime(year, 12, 31).date() if month == 12 else (
                datetime(year, month + 1, 1).date() - timedelta(days=1))
    else:
        start_date = _parse_date(request.args.get('start'))
        end_date = _parse_date(request.args.get('end'))
        if not (start_date and end_date):
            return jsonify({"error": "Provide start and end in YYYY-MM-DD for custom scope"}), 400

    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    q = (invoice.query
         .options(joinedload(invoice.customer))
         .join(customer, invoice.customerId == customer.id)
         .filter(invoice.isDeleted == False,
                 customer.isDeleted == False,
                 invoice.createdAt >= start_dt,
                 invoice.createdAt <= end_dt))

    if phone:
        q = q.filter(customer.phone == phone)
    invs = q.order_by(invoice.createdAt.asc()).all()

    totals = {
        "invoice_count": len(invs),
        "amount": round(sum((inv.totalAmount or 0) for inv in invs), 2)
    }

    per_customer = defaultdict(lambda: {"count": 0, "amount": 0.0})
    per_day = defaultdict(lambda: {"count": 0, "amount": 0.0})
    per_month = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for inv in invs:
        cust = inv.customer
        cust_key = f"{cust.name} ({cust.phone})" if cust else "Unknown"
        per_customer[cust_key]["count"] += 1
        per_customer[cust_key]["amount"] += (inv.totalAmount or 0)
        dkey = inv.createdAt.strftime('%Y-%m-%d')
        per_day[dkey]["count"] += 1
        per_day[dkey]["amount"] += (inv.totalAmount or 0)
        mkey = inv.createdAt.strftime('%Y-%m')
        per_month[mkey]["count"] += 1
        per_month[mkey]["amount"] += (inv.totalAmount or 0)

    # Convert defaultdicts to plain dicts for JSON
    return jsonify({
        "range": {"start": start_date.isoformat(), "end": end_date.isoformat()},
        "totals": totals,
        "per_customer": {k: {"count": v["count"], "amount": round(v["amount"], 2)} for k, v in per_customer.items()},
        "per_day": {k: {"count": v["count"], "amount": round(v["amount"], 2)} for k, v in per_day.items()},
        "per_month": {k: {"count": v["count"], "amount": round(v["amount"], 2)} for k, v in per_month.items()},
    })


@app.route('/api/statements/invoices', methods=['GET'])
def api_statements_invoices():
    """JSON: raw invoice rows with pagination (for tables/exports).
    Query params:
      scope/year/month/start/end/phone (same as above)
      page (default 1), per_page (default 50, max 500)
    """
    scope = (request.args.get('scope') or 'custom').lower()
    phone = request.args.get('phone')
    page = max(int(request.args.get('page', 1)), 1)
    per_page = min(max(int(request.args.get('per_page', 50)), 1), 500)

    today = datetime.now().date()
    if scope == 'year':
        year = int(request.args.get('year') or today.year)
        start_date = datetime(year, 1, 1).date()
        end_date = datetime(year, 12, 31).date()
    elif scope == 'month':
        year = int(request.args.get('year') or today.year)
        month = int(request.args.get('month') or today.month)
        start_date = datetime(year, month, 1).date()
        end_date = datetime(year, 12, 31).date() if month == 12 else (
                datetime(year, month + 1, 1).date() - timedelta(days=1))
    else:
        start_date = _parse_date(request.args.get('start'))
        end_date = _parse_date(request.args.get('end'))
        if not (start_date and end_date):
            return jsonify({"error": "Provide start and end in YYYY-MM-DD for custom scope"}), 400

    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    q = (invoice.query
         .options(joinedload(invoice.customer))
         .join(customer, invoice.customerId == customer.id)
         .filter(invoice.isDeleted == False,
                 customer.isDeleted == False,
                 invoice.createdAt >= start_dt,
                 invoice.createdAt <= end_dt))

    if phone:
        q = q.filter(customer.phone == phone)

    total = q.count()
    invs = q.order_by(invoice.createdAt.asc()).offset((page - 1) * per_page).limit(per_page).all()

    rows = []
    for inv in invs:
        cust = inv.customer
        rows.append({
            "invoice_no": inv.invoiceId,
            "date": inv.createdAt.strftime('%Y-%m-%d'),
            "customer": cust.name if cust else 'Unknown',
            "phone": cust.phone if cust else '',
            "total": round(inv.totalAmount or 0, 2)
        })

    return jsonify({
        "range": {"start": start_date.isoformat(), "end": end_date.isoformat()},
        "total": total,
        "page": page,
        "per_page": per_page,
        "rows": rows
    })


@app.route('/statements/blank', methods=['GET'])
def statements_blank():
    return render_template(
        'statement.html',
        start_date=None,
        end_date=None,
        total_invoices=0,
        total_amount=0,
        phone=None,
        per_customer={},
        invs=[],
        scope='custom',
    )


@app.route('/create-bill', methods=['GET', 'POST'])
def start_bill():
    # --- GET: prefill if customer_id is supplied; otherwise show selector ---
    if request.method == 'GET':
        cid = request.args.get('customer_id', type=int)
        if cid:
            try:
                cust = (customer.query
                        .filter(customer.isDeleted == False, customer.id == cid)
                        .first())
            except Exception:
                cust = None
            if not cust:
                flash('Customer not found', 'warning')
                return redirect(url_for('about_user'))
            inventory_list = item.query.order_by(item.name.asc()).all()
            return _render_create_bill(customer=cust, inventory=inventory_list)
        # GET: no customer_id, just render blank/new bill
        return _render_create_bill()

    # POST logic
    # POST (A) select customer
    if 'description[]' not in request.form:
        selected_phone = request.form.get("customer") or request.form.get("customer_phone")
        sel = (customer.query
               .filter(customer.isDeleted == False, customer.phone == selected_phone)
               .first())
        if not sel:
            flash('Please pick a valid customer', 'warning')
            return render_template('select_customer.html')
        return _render_create_bill(customer=sel, inventory=item.query.all())

    # (B) Final bill submission with line items
    submitted_token = request.form.get('form_token')
    if not _validate_bill_token(submitted_token):
        flash('The bill form has expired. Please start a new bill.', 'warning')
        return redirect(url_for('select_customer'))

    selected_phone = request.form.get('customer_phone')
    selected_customer = customer.query.filter_by(phone=selected_phone).first()
    if not selected_customer:
        flash('Customer not found. Please reselect the customer.', 'warning')
        return render_template('select_customer.html')

    descriptions = request.form.getlist('description[]')
    quantities = request.form.getlist('quantity[]')
    rates = request.form.getlist('rate[]')
    dc_numbers = request.form.getlist('dc_no[]')  # may be [] if toggle off
    rounded_flags = request.form.getlist('rounded[]')

    # Get exclusion flags
    exclude_phone = bool(request.form.get('exclude_phone'))
    exclude_gst = bool(request.form.get('exclude_gst'))
    exclude_addr = bool(request.form.get('exclude_addr'))

    total = 0.0
    item_rows = []
    for i in range(len(descriptions)):
        desc = (descriptions[i] or '').strip()
        if not desc:
            continue
        qty = int(quantities[i]) if i < len(quantities) and quantities[i] else 0
        rate = float(rates[i]) if i < len(rates) and rates[i] else 0.0
        dc_val = ''
        if dc_numbers and i < len(dc_numbers) and dc_numbers[i]:
            dc_val = dc_numbers[i].strip()
        rounded = (i < len(rounded_flags) and rounded_flags[i] == "1")
        raw_total = qty * rate
        line_total = rounding_to_nearest_zero(raw_total) if rounded else raw_total
        total += line_total
        item_rows.append([desc, qty, rate, line_total, dc_val])

    # Create invoice
    new_invoice = invoice(
        customerId=selected_customer.id,
        createdAt=datetime.now(timezone.utc),
        totalAmount=(round(total, 2)),
        pdfPath="",  # set after inv_name built
        invoiceId="",  # temporary
        exclude_phone=exclude_phone,
        exclude_gst=exclude_gst,
        exclude_addr=exclude_addr
    )

    db.session.add(new_invoice)
    db.session.flush()
    # Add Alert - Not needed

    # Generate invoice Id + pdf path
    inv_name = f"SLP-{datetime.now().strftime('%d%m%y')}-{str(new_invoice.id).zfill(5)}"
    pdf_filename = f"{inv_name}.pdf"
    pdf_path = os.path.join("static/pdfs", pdf_filename)

    new_invoice.invoiceId = inv_name
    new_invoice.pdfPath = pdf_path
    db.session.flush()
    # add alerts - not needed as persistant on in place

    # Add line items
    for desc, qty, rate, line_total, dc_val in item_rows:
        matched_item = item.query.filter_by(name=desc).first()
        if matched_item:
            item_id = matched_item.id
        else:
            new_item = item(name=desc, unitPrice=rate, quantity=0, taxPercentage=0)
            db.session.add(new_item)
            db.session.flush()
            # add alert - not needed as persistent one in place
            item_id = new_item.id

        db.session.add(invoiceItem(
            invoiceId=new_invoice.id,
            itemId=item_id,
            quantity=qty,
            rate=rate,
            discount=0,
            taxPercentage=0,
            line_total=line_total,
            dcNo=(dc_val if dc_val else None)
        ))
    db.session.commit()

    # add alerts - Not needed, persistent one is in place

    # Did user include any DC values?
    # dc_present = any((x or '').strip() for x in (dc_numbers or []))

    # After successful creation, flash and redirect to locked preview page
    session['persistent_notice'] = f"Invoice {new_invoice.invoiceId} created successfully!"
    return redirect(url_for('view_bill_locked', invoicenumber=new_invoice.invoiceId, edit_bill='true'))


@app.route('/view_customers', methods=['GET', 'POST'])
def view_customers():
    if request.method == 'POST':
        phone = request.form.get('customer')
        sel = (customer.query
               .filter(customer.isDeleted == False, customer.phone == phone)
               .first_or_404())
        return _render_create_bill(customer=sel, inventory=item.query.all())

    query = (request.args.get('q') or '').strip().lower()
    q = (customer.query
         .filter(customer.isDeleted == False)
         .order_by(customer.createdAt.desc(), customer.id.desc()))
    customers = q.all()

    if query:
        customers = [
            c for c in customers
            if query in (c.company or '').lower()
               or query in (c.name or '').lower()
               or query in (c.phone or '')
        ]

    return render_template('view_customers.html', customers=customers)


@app.route('/view_bills')
def view_bills():
    """Render all bills with filtering, search, and sorting."""

    # ---- 1️⃣ Extract filters from query params ----
    query = (request.args.get('q') or '').lower()
    phone = request.args.get('phone')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # ---- 2️⃣ Base query with eager loading ----
    q = (
        invoice.query
        .options(joinedload(invoice.customer))
        .join(customer, invoice.customerId == customer.id)
        .filter(invoice.isDeleted == False, customer.isDeleted == False)
    )

    # ---- 3️⃣ Sorting ----
    sort_key = (request.args.get('sort') or 'date').lower()
    sort_dir = (request.args.get('dir') or 'desc').lower()

    def order(col):
        return col.desc() if sort_dir == 'desc' else col.asc()

    if sort_key == 'total':
        q = q.order_by(order(invoice.totalAmount))
    elif sort_key == 'invoice':
        q = q.order_by(order(invoice.invoiceId))
    elif sort_key == 'customer':
        q = q.order_by(order(customer.name))
    else:
        q = q.order_by(order(invoice.createdAt))

    # ---- 4️⃣ Optional date range filter ----
    try:
        if start_date and end_date:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            q = q.filter(invoice.createdAt >= start_dt, invoice.createdAt < end_dt)
    except Exception:
        pass

    # ---- 5️⃣ Execute main query ----
    invoices = q.all()

    # ---- 6️⃣ Transform for template ----
    bills = []
    for inv in invoices:
        cust = inv.customer
        bills.append({
            "invoice_no": inv.invoiceId,
            "date": inv.createdAt.strftime('%d-%b-%Y'),
            "customer_name": cust.name if cust else 'Unknown',
            "phone": cust.phone if cust else '',
            "total": f"{inv.totalAmount:,.2f}",
            "filename": f"{inv.invoiceId}.pdf",
            "customer_company": cust.company if cust else 'Unknown',
            "is_paid": bool(getattr(inv, 'payment', False))
        })

    # ---- 7️⃣ Apply search filters ----
    if phone:
        bills = [b for b in bills if b['phone'] == phone]
    elif query:
        bills = [
            b for b in bills
            if query in b['customer_name'].lower()
               or query in b.get('phone', '')
               or query in b['invoice_no'].lower()
               or query in (b.get('customer_company') or '').lower()
        ]

    # ---- 8️⃣ Render ----
    current_filters_url = request.full_path if request.query_string else request.path
    return render_template('view_bills.html', bills=bills, mark_paid_redirect=current_filters_url)


@app.route('/view-bill/<invoicenumber>')
def view_bill_locked(invoicenumber):
    # load invoice and related data
    current_invoice = invoice.query.filter_by(invoiceId=invoicenumber, isDeleted=False).first_or_404()
    cur_cust = customer.query.get(current_invoice.customerId)
    line_items = invoiceItem.query.filter_by(invoiceId=current_invoice.id).all()

    current_customer = {
        "name": cur_cust.name,
        "company": cur_cust.company,
        "phone": "Excluded in the bill" if current_invoice.exclude_phone else cur_cust.phone,
        "gst": "Excluded in the bill" if current_invoice.exclude_gst else cur_cust.gst,
        "address": "Excluded in the bill" if current_invoice.exclude_addr else cur_cust.address,
        "email": cur_cust.email
    }

    # build row wise lists for the template
    descriptions, quantities, rates, dc_numbers = [], [], [], []
    line_totals = []

    total = 0.0
    for li in line_items:
        itm = item.query.get(li.itemId)
        descriptions.append(itm.name if itm else 'Unknown')
        quantities.append(li.quantity)
        rates.append(li.rate)
        dc_numbers.append(li.dcNo or '')
        line_totals.append(li.line_total)
        total += li.line_total or 0

    # Determine whether to show DC column
    dcno = any((x or '').strip() for x in dc_numbers)

    edit_bill = request.args.get('edit_bill', '').lower() in ('yes', 'true', '1')
    back_two_pages = edit_bill

    invoice_date = current_invoice.createdAt
    invoice_paid = bool(getattr(current_invoice, 'payment', False))

    current_page_url = request.full_path if request.query_string else request.path
    return render_template(
        'view_bill_locked.html',
        customer=current_customer,
        descriptions=descriptions,
        quantities=quantities,
        rates=rates,
        dc_numbers=dc_numbers,
        line_totals=line_totals,
        dcno=dcno,
        total=round(total, 2),
        invoice_no=current_invoice.invoiceId,
        back_to_select_customer=False,
        back_to_url=None,
        customer_id=cur_cust.id,
        back_two_pages=back_two_pages,
        invoice_date=invoice_date,
        mark_paid_redirect=current_page_url,
        is_paid=invoice_paid,
    )


def mark_bill_paid(invoice_no):
    invoice_obj = invoice.query.filter_by(invoiceId=invoice_no, isDeleted=False).first_or_404()
    customer_obj = customer.query.filter_by(id=invoice_obj.customerId, isDeleted=False).first()

    raw_next = request.form.get('next') or ''
    next_url = raw_next or url_for('view_bills')
    parsed_next = urlparse(next_url)
    if parsed_next.netloc and parsed_next.netloc != request.host:
        next_url = url_for('view_bills')

    if getattr(invoice_obj, 'payment', False):
        flash('Invoice already marked as paid.', 'info')
        return redirect(next_url)

    if not customer_obj:
        flash('Unable to record payment: customer details missing for this invoice.', 'danger')
        return redirect(next_url)

    amount_value = float(invoice_obj.totalAmount or 0.0)
    if amount_value <= 0:
        flash('Unable to record payment for an empty invoice.', 'warning')
        return redirect(next_url)

    source = (request.form.get('source') or 'view_bills').strip().lower()
    remarks_map = {
        'view_bill_locked': 'Marked as paid via bill detail page.',
        'view_bills': 'Marked as paid via bills list.',
    }
    remarks = remarks_map.get(source, 'Marked as paid.')

    txn = accountingTransaction(
        customerId=customer_obj.id,
        amount=amount_value,
        txn_type='income',
        mode='bank',
        account='current',
        invoice_no=invoice_obj.invoiceId,
        remarks=remarks,
    )
    db.session.add(txn)
    db.session.flush()
    _sync_invoice_payment_flag(invoice_obj.invoiceId)
    try:
        db.session.commit()
        flash(f'Recorded payment of INR {amount_value:,.2f} for invoice {invoice_obj.invoiceId}.', 'success')
    except Exception as exc:
        db.session.rollback()
        flash(f'Unable to record payment: {exc}', 'danger')

    return redirect(next_url)


app.add_url_rule(
    '/bills/<invoice_no>/mark-paid',
    view_func=mark_bill_paid,
    endpoint='mark_bill_paid',
    methods=['POST']
)


# amounts to words
# --- Amount to words (Indian numbering: Crore, Lakh, Thousand) ---
ONES = [
    "", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
    "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
    "Seventeen", "Eighteen", "Nineteen"
]
TENS = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]


def _two_digits(n: int) -> str:
    if n == 0:
        return ""
    if n < 20:
        return ONES[n]
    return (TENS[n // 10] + (" " + ONES[n % 10] if (n % 10) != 0 else "")).strip()


def _three_digits(n: int) -> str:
    # 0..999
    if n >= 100:
        rem = n % 100
        s = ONES[n // 100] + " Hundred"
        if rem:
            s += " and " + _two_digits(rem)
        return s
    return _two_digits(n)


def rupees_to_words(num: int) -> str:
    if num == 0:
        return "Zero"
    parts = []
    crore = num // 10000000;
    num %= 10000000
    lakh = num // 100000;
    num %= 100000
    thousand = num // 1000;
    num %= 1000
    rest = num  # 0..999

    # Use _three_digits for groups that can be up to 999
    if crore:
        parts.append(_three_digits(int(crore)) + " Crore")
    if lakh:
        parts.append(_three_digits(int(lakh)) + " Lakh")
    if thousand:
        parts.append(_three_digits(int(thousand)) + " Thousand")
    if rest:
        parts.append(_three_digits(int(rest)))

    return " ".join(parts)


def amount_to_words(amount) -> str:
    try:
        amt = float(amount or 0)
    except Exception:
        amt = 0.0
    rupees = int(amt)
    paise = int(round((amt - rupees) * 100))
    words = rupees_to_words(rupees) + " Rupees"
    if paise:
        words += " and " + _two_digits(paise) + " Paise"
    return words + " Only"


@app.route('/bill_preview/<invoicenumber>')
def bill_preview(invoicenumber):
    current_invoice = invoice.query.filter_by(invoiceId=invoicenumber, isDeleted=False).first_or_404()
    if not current_invoice:
        return f"No invoice found for {invoicenumber}"

    cur_cust = customer.query.get(current_invoice.customerId)

    current_customer = {
        "name": cur_cust.name,
        "company": cur_cust.company,
        "phone": None if current_invoice.exclude_phone else cur_cust.phone,
        "gst": None if current_invoice.exclude_gst else cur_cust.gst,
        "address": None if current_invoice.exclude_addr else cur_cust.address,
        "email": cur_cust.email
    }
    items = invoiceItem.query.filter_by(invoiceId=current_invoice.id).all()

    # Prepare item data
    item_data = []
    for i in items:
        item_name = item.query.get(i.itemId).name if i.itemId else "Unknown"
        entry = (
            item_name,
            "N/A",
            i.quantity,
            i.rate,
            i.discount,
            i.taxPercentage,
            i.line_total
        )
        item_data.append(entry)

    # DC numbers
    dc_numbers = [i.dcNo or '' for i in items]
    dcno = any(bool((x or '').strip()) for x in dc_numbers)

    config = layoutConfig().get_or_create()
    current_sizes = config.get_sizes()

    upi_id = APP_INFO["upi_info"]["upi_id"]
    company_name = APP_INFO["business"]["name"]
    upi_name = APP_INFO["upi_info"]["upi_name"]
    brand_watermark_path = _resolve_brand_watermark_path(APP_INFO.get("business"))
    brand_accent_color = _resolve_brand_accent_color(APP_INFO.get("business"))

    api_url = f"{request.host_url.rstrip('/')}/api/generate_upi_qr"
    params = {"upi_id": upi_id, "amount": current_invoice.totalAmount, "company_name": company_name}

    try:
        resp = requests.get(api_url, params=params, timeout=5)
        if resp.status_code == 200:
            data = resp.json()
            qr_svg_base64 = data.get('qr_svg_base64')
            upi_url = data.get('upi_url')
        else:
            qr_svg_base64 = None
            upi_url = None
    except Exception as e:
        print(f"[ERROR] failed to fetch QR: {e}")
        qr_svg_base64 = None
        upi_url = None

    return render_template(
        'bill_preview.html',
        invoice=current_invoice,
        customer=current_customer,
        items=item_data,
        dcno=dcno,
        dc_numbers=dc_numbers,
        total_in_words=amount_to_words(current_invoice.totalAmount),
        sizes=current_sizes,
        qr_svg_base64=qr_svg_base64,
        upi_id=upi_id,
        upi_name=upi_name,
        company_name=company_name,
        total=current_invoice.totalAmount,
        app_info=APP_INFO,
        brand_watermark_path=brand_watermark_path,
        brand_accent_color=brand_accent_color,
    )


@app.route('/edit-bill/<invoicenumber>', methods=['GET', 'POST'])
def edit_bill(invoicenumber):
    # fetch invoice and related data
    current_invoice = invoice.query.filter_by(invoiceId=invoicenumber).first_or_404()
    current_customer = customer.query.get(current_invoice.customerId)
    line_items = invoiceItem.query.filter_by(invoiceId=current_invoice.id).all()

    # Build lists for template
    descriptions, quantities, rates, dc_numbers = [], [], [], []
    line_totals = []
    total = 0.0
    for li in line_items:
        itm = item.query.get(li.itemId)
        descriptions.append(itm.name if itm else 'Unknown')
        quantities.append(li.quantity)
        rates.append(li.rate)
        dc_numbers.append(li.dcNo or '')
        line_totals.append(li.line_total or 0)
        total += li.line_total or 0

    dcno = any((x or '').strip() for x in dc_numbers)

    prev_invoice_no = current_invoice.invoiceId
    try:
        prev_created_at = current_invoice.createdAt.strftime('%Y-%m-%d %H:%M')
    except Exception:
        prev_created_at = str(current_invoice.createdAt)

    exclude_phone = current_invoice.exclude_phone
    exclude_gst = current_invoice.exclude_gst
    exclude_addr = current_invoice.exclude_addr

    # If POST: update invoice and redirect to view_bill_locked
    if request.method == 'POST':
        # Update customer-level metadata before saving invoice
        current_invoice.exclude_phone = request.form.get('exclude_phone') in ('on', 'true', '1')
        current_invoice.exclude_gst = request.form.get('exclude_gst') in ('on', 'true', '1')
        current_invoice.exclude_addr = request.form.get('exclude_addr') in ('on', 'true', '1')
        db.session.commit()
        # add alert - Not needed funcionally
        return redirect(url_for('view_bill_locked', invoicenumber=current_invoice.invoiceId, edit_bill='true'))

    # Render the same template as create_bill.html but pre-filled
    return _render_create_bill(
        customer=current_customer,
        inventory=item.query.all(),
        success=False,  # show filled rows
        descriptions=descriptions,
        quantities=quantities,
        rates=rates,
        dc_numbers=dc_numbers,
        dcno=dcno,
        total=round(total, 2),
        grand_total=round(total, 2),
        invoice_no=current_invoice.invoiceId,
        edit_mode=True,  # flag to distinguish editing vs new bill
        prev_invoice_no=prev_invoice_no,
        prev_created_at=prev_created_at,
        exclude_phone=exclude_phone,
        exclude_gst=exclude_gst,
        exclude_addr=exclude_addr,
        line_totals=line_totals,
    )


@app.route('/delete-bill/<invoicenumber>', methods=['POST'])
def delete_bill(invoicenumber):
    inv = invoice.query.filter_by(invoiceId=invoicenumber, isDeleted=False).first_or_404()
    inv.isDeleted = True
    inv.deletedAt = datetime.now(timezone.utc)
    db.session.commit()
    # add alert
    flash('Bill has been deleted.', 'danger')

    next_url = request.form.get('next') or ''
    try:
        host = urlparse(request.host_url).netloc
        parsed = urlparse(next_url)
        if next_url and (parsed.netloc == '' or parsed.netloc == host):
            return redirect(next_url)
    except Exception:
        pass
    return redirect(url_for('view_bills'))


@app.route('/update-bill/<invoicenumber>', methods=['POST'])
def update_bill(invoicenumber):
    # 1) Load the invoice being edited
    current_invoice = invoice.query.filter_by(invoiceId=invoicenumber, isDeleted=False).first_or_404()
    current_customer = customer.query.get(current_invoice.customerId)

    submitted_token = request.form.get('form_token')
    if not _validate_bill_token(submitted_token):
        flash('The bill form has expired. Please reopen the invoice before updating.', 'warning')
        return redirect(url_for('view_bill_locked', invoicenumber=invoicenumber, edit_bill='true'))

    # 2) Read form inputs
    descriptions = request.form.getlist('description[]')
    quantities = request.form.getlist('quantity[]')
    rates = request.form.getlist('rate[]')
    dc_numbers = request.form.getlist('dc_no[]')  # may be empty if toggle off
    rounded_flags = request.form.getlist('rounded[]')

    # 3) Normalize rows + recompute totals
    rows = []
    total = 0.0
    for i in range(len(descriptions)):
        desc = (descriptions[i] or '').strip()
        if not desc:
            continue  # skip empty rows

        qty = int(quantities[i]) if i < len(quantities) and quantities[i] else 0
        rate = float(rates[i]) if i < len(rates) and rates[i] else 0.0
        dc = (dc_numbers[i].strip() if i < len(dc_numbers) and dc_numbers[i] else None)
        rounded = (i < len(rounded_flags) and rounded_flags[i] == "1")
        raw_total = qty * rate
        line_total = rounding_to_nearest_zero(raw_total) if rounded else raw_total
        total += line_total
        rows.append((desc, qty, rate, dc, line_total))

    # 4) Replace all existing line items with the new set using ORM deletes so sync events fire
    existing_items = invoiceItem.query.filter_by(invoiceId=current_invoice.id).all()
    for existing_item in existing_items:
        db.session.delete(existing_item)
    db.session.flush()

    for desc, qty, rate, dc, line_total in rows:
        # Reuse existing item by name, or create a placeholder item if not found
        matched_item = item.query.filter_by(name=desc).first()
        if matched_item:
            item_id = matched_item.id
        else:
            new_item = item(name=desc, unitPrice=rate, quantity=0, taxPercentage=0)
            db.session.add(new_item)
            db.session.flush()
            item_id = new_item.id

        db.session.add(invoiceItem(
            invoiceId=current_invoice.id,
            itemId=item_id,
            quantity=qty,
            rate=rate,
            discount=0,
            taxPercentage=0,
            line_total=line_total,
            dcNo=dc if dc else None
        ))

    # 5) Update invoice total (and updatedAt if you have it)
    current_invoice.totalAmount = (round(total, 2))

    # 5.5) Update customer-level metadata before saving invoice
    current_invoice.exclude_phone = request.form.get('exclude_phone') in ('on', 'true', '1')
    current_invoice.exclude_gst = request.form.get('exclude_gst') in ('on', 'true', '1')
    current_invoice.exclude_addr = request.form.get('exclude_addr') in ('on', 'true', '1')

    db.session.commit()
    # add alert - not needed persistent one in place

    # 6) Redirect to locked preview after update
    session['persistent_notice'] = f"Old invoice {current_invoice.invoiceId} updated successfully!"

    return redirect(url_for('view_bill_locked', invoicenumber=current_invoice.invoiceId, edit_bill='true', new_bill='true'))


@app.route('/bill_preview/latest')
def latest_bill_preview():
    current_invoice = (
        invoice.query
        .filter(invoice.isDeleted == False)
        .order_by(invoice.id.desc())
        .first()
    )
    if not current_invoice:
        return "No invoice found"

    cur_cust = customer.query.get(current_invoice.customerId)

    current_customer = {
        "name": cur_cust.name,
        "company": cur_cust.company,
        "phone": None if current_invoice.exclude_phone else cur_cust.phone,
        "gst": None if current_invoice.exclude_gst else cur_cust.gst,
        "address": None if current_invoice.exclude_addr else cur_cust.address,
        "email": cur_cust.email
    }

    items = invoiceItem.query.filter_by(invoiceId=current_invoice.id).all()
    item_data = []

    for i in items:
        item_name = item.query.get(i.itemId).name if i.itemId else "Unknown"
        entry = (
            item_name,
            "N/A",
            i.quantity,
            i.rate,
            i.discount,
            i.taxPercentage,
            i.line_total
        )
        item_data.append(entry)

    config = layoutConfig.get_or_create()
    current_sizes = config.get_sizes()

    dc_numbers = [i.dcNo or '' for i in items]
    dcno = any(bool((x or '').strip()) for x in dc_numbers)

    # 🔹 UPI QR generation (same as main bill_preview route)
    upi_id = APP_INFO["upi_info"]["upi_id"]
    company_name = APP_INFO["business"]["name"]
    upi_name = APP_INFO["upi_info"]["upi_name"]
    brand_watermark_path = _resolve_brand_watermark_path(APP_INFO.get("business"))
    brand_accent_color = _resolve_brand_accent_color(APP_INFO.get("business"))

    api_url = f"{request.host_url.rstrip('/')}/api/generate_upi_qr"
    params = {
        "upi_id": upi_id,
        "amount": current_invoice.totalAmount,
        "company_name": company_name
    }

    try:
        resp = requests.get(api_url, params=params, timeout=5)
        if resp.status_code == 200:
            data = resp.json()
            qr_svg_base64 = data.get('qr_svg_base64')
            upi_url = data.get('upi_url')
        else:
            qr_svg_base64 = None
            upi_url = None
    except Exception as e:
        print(f"[ERROR] failed to fetch QR: {e}")
        qr_svg_base64 = None
        upi_url = None

    return render_template(
        'bill_preview.html',
        invoice=current_invoice,
        customer=current_customer,
        items=item_data,
        dcno=dcno,
        dc_numbers=dc_numbers,
        total_in_words=amount_to_words(current_invoice.totalAmount),
        sizes=current_sizes,
        app_info=APP_INFO,
        qr_svg_base64=qr_svg_base64,
        upi_id=upi_id,
        upi_name=upi_name,
        company_name=company_name,
        total=current_invoice.totalAmount,
        brand_watermark_path=brand_watermark_path,
        brand_accent_color=brand_accent_color,
    )


app.jinja_env.globals.update(zip=zip)


@app.route('/statements', methods=['GET'])
def statements():
    """
    Render customer statements for a date range or customer, with export (HTML, CSV, PDF/Print).
    Query params:
      - scope: 'year'/'month'/'custom'
      - year/month/start/end
      - phone: filter by customer phone
      - export: 'csv' or 'pdf' (default: html)
    """
    scope = (request.args.get('scope') or 'custom').lower()
    phone = request.args.get('phone')
    export = (request.args.get('export') or '').lower()

    today = datetime.now().date()
    min_allowed_start = get_default_statement_start().date()  # 🔹 Lower limit from info.json

    # 🔸 Resolve start/end based on scope
    if scope == 'year':
        year = int(request.args.get('year') or today.year)
        start_date = datetime(year, 1, 1).date()
        end_date = datetime(year, 12, 31).date()
    elif scope == 'month':
        year = int(request.args.get('year') or today.year)
        month = int(request.args.get('month') or today.month)
        start_date = datetime(year, month, 1).date()
        end_date = datetime(year, 12, 31).date() if month == 12 else (
                datetime(year, month + 1, 1).date() - timedelta(days=1))
    else:
        start_date = _parse_date(request.args.get('start'))
        end_date = _parse_date(request.args.get('end'))
        # Fallback: default to current month
        if not (start_date and end_date):
            start_date = today.replace(day=1)
            end_date = today

    # 🔸 Enforce lower date limit
    if start_date < min_allowed_start:
        start_date = min_allowed_start

    # Normalize datetimes
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    # 🔹 Query invoices
    q = (
        invoice.query
        .options(joinedload(invoice.customer))
        .join(customer, invoice.customerId == customer.id)
        .filter(
            invoice.isDeleted == False,
            customer.isDeleted == False,
            invoice.createdAt >= start_dt,
            invoice.createdAt <= end_dt,
        )
    )
    if phone:
        q = q.filter(customer.phone == phone)

    invs = q.order_by(invoice.createdAt.asc()).all()

    # 🔹 Aggregations
    total_invoices = len(invs)
    total_amount = round(sum((inv.totalAmount or 0) for inv in invs), 2)
    per_customer = defaultdict(lambda: {"count": 0, "amount": 0.0, "company": None, "phone": None})

    for inv in invs:
        cust = inv.customer
        if cust:
            key = cust.phone
            per_customer[key]["count"] += 1
            per_customer[key]["amount"] += (inv.totalAmount or 0)
            per_customer[key]["company"] = cust.company
            per_customer[key]["phone"] = cust.phone
        else:
            per_customer["Unknown"]["count"] += 1
            per_customer["Unknown"]["amount"] += (inv.totalAmount or 0)

    if not per_customer:
        per_customer = {}

    # 🔸 Export CSV
    if export == "csv":
        output = io.StringIO()
        writer = csv.writer(output)

        biz = APP_INFO.get("business", {})
        bank = APP_INFO.get("bank", {})
        statement_meta = APP_INFO.get("statement", {})

        writer.writerow(
            [f"{biz.get('name', 'Business Name')} - {statement_meta.get('header_title', 'Statement Summary')}"])
        writer.writerow(["Generated On", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow(["Period", f"{start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}"])
        writer.writerow([])

        writer.writerow(["Invoice No", "Date", "Company", "Phone", "Amount (INR)"])
        for inv in invs:
            cust = inv.customer
            writer.writerow([
                inv.invoiceId,
                inv.createdAt.strftime('%d %B %Y'),
                cust.company if cust else '',
                cust.phone if cust else '',
                f"{inv.totalAmount:,.2f}"
            ])

        writer.writerow([])
        writer.writerow(["Summary"])
        writer.writerow(["Total Invoices", total_invoices])
        writer.writerow(["Total Amount (INR)", f"{total_amount:,.2f}"])
        writer.writerow([])

        if per_customer:
            writer.writerow(["Per Customer Summary"])
            writer.writerow(["Phone", "Company", "Invoice Count", "Total Amount (INR)"])
            for key, val in per_customer.items():
                writer.writerow([
                    val.get("phone", key),
                    val.get("company", ""),
                    val.get("count", 0),
                    f"{val.get('amount', 0):,.2f}"
                ])
            writer.writerow([])

        writer.writerow(["Payment Information"])
        writer.writerow(["Account Name", bank.get("account_name", "")])
        writer.writerow(["Bank", f"{bank.get('bank_name', '')}, {bank.get('branch', '')}"])
        writer.writerow(["Account Number", bank.get("account_number", "")])
        writer.writerow(["IFSC", bank.get("ifsc", "")])
        writer.writerow(["PhonePe/GPay", biz.get("upi_id", "")])
        writer.writerow([])

        writer.writerow(["Disclaimer", statement_meta.get("disclaimer", "This is a system-generated statement.")])

        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename=statements_{start_date}_{end_date}.csv"}
        )

    # 🔸 Export XLSX
    if export == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, numbers
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws_inv = wb.active
        ws_inv.title = "Invoices"

        # Header row
        headers = ["Invoice No", "Date", "Company", "Phone", "Amount (INR)"]
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        for col_num, col_name in enumerate(headers, 1):
            cell = ws_inv.cell(row=1, column=col_num, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Data rows
        currency_fmt = u'INR #,##0.00'
        for row_num, inv in enumerate(invs, 2):
            cust = inv.customer
            ws_inv.cell(row=row_num, column=1, value=inv.invoiceId)
            ws_inv.cell(row=row_num, column=2, value=inv.createdAt.strftime('%d %B %Y'))
            ws_inv.cell(row=row_num, column=3, value=cust.company if cust else '')
            ws_inv.cell(row=row_num, column=4, value=cust.phone if cust else '')
            amt_cell = ws_inv.cell(row=row_num, column=5, value=round(inv.totalAmount or 0, 2))
            amt_cell.number_format = currency_fmt
            amt_cell.alignment = Alignment(horizontal="right")

        # Auto-size columns for "Invoices"
        for col in ws_inv.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(val))
                except Exception:
                    pass
            ws_inv.column_dimensions[col_letter].width = min(max_length + 2, 40)

        # Add "Summary" sheet
        ws_sum = wb.create_sheet("Summary")
        row = 1
        bold = Font(bold=True)
        # Title
        biz = APP_INFO.get("business", {})
        bank = APP_INFO.get("bank", {})
        statement_meta = APP_INFO.get("statement", {})
        ws_sum.cell(row=row, column=1,
                    value=f"{biz.get('name', 'Business Name')} - {statement_meta.get('header_title', 'Statement Summary')}").font = bold
        row += 1
        ws_sum.cell(row=row, column=1, value="Generated On")
        ws_sum.cell(row=row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        row += 1
        ws_sum.cell(row=row, column=1, value="Period")
        ws_sum.cell(row=row, column=2, value=f"{start_date.strftime('%d %B %Y')} to {end_date.strftime('%d %B %Y')}")
        row += 2
        ws_sum.cell(row=row, column=1, value="Summary").font = bold
        row += 1
        ws_sum.cell(row=row, column=1, value="Total Invoices")
        ws_sum.cell(row=row, column=2, value=total_invoices)
        row += 1
        ws_sum.cell(row=row, column=1, value="Total Amount (INR)")
        amt_cell = ws_sum.cell(row=row, column=2, value=total_amount)
        amt_cell.number_format = currency_fmt
        row += 2

        # Per-customer summary
        if per_customer:
            ws_sum.cell(row=row, column=1, value="Per Customer Summary").font = bold
            row += 1
            ws_sum.cell(row=row, column=1, value="Phone").font = bold
            ws_sum.cell(row=row, column=2, value="Company").font = bold
            ws_sum.cell(row=row, column=3, value="Invoice Count").font = bold
            ws_sum.cell(row=row, column=4, value="Total Amount (INR)").font = bold
            row += 1
            for key, val in per_customer.items():
                ws_sum.cell(row=row, column=1, value=val.get("phone", key))
                ws_sum.cell(row=row, column=2, value=val.get("company", ""))
                ws_sum.cell(row=row, column=3, value=val.get("count", 0))
                amt_cell = ws_sum.cell(row=row, column=4, value=round(val.get("amount", 0), 2))
                amt_cell.number_format = currency_fmt
                row += 1
            row += 1

        # Payment Information
        ws_sum.cell(row=row, column=1, value="Payment Information").font = bold
        row += 1
        ws_sum.cell(row=row, column=1, value="Account Name")
        ws_sum.cell(row=row, column=2, value=bank.get("account_name", ""))
        row += 1
        ws_sum.cell(row=row, column=1, value="Bank")
        ws_sum.cell(row=row, column=2, value=f"{bank.get('bank_name', '')}, {bank.get('branch', '')}")
        row += 1
        ws_sum.cell(row=row, column=1, value="Account Number")
        ws_sum.cell(row=row, column=2, value=bank.get("account_number", ""))
        row += 1
        ws_sum.cell(row=row, column=1, value="IFSC")
        ws_sum.cell(row=row, column=2, value=bank.get("ifsc", ""))
        row += 1
        ws_sum.cell(row=row, column=1, value="PhonePe/GPay")
        ws_sum.cell(row=row, column=2, value=biz.get("upi_id", ""))
        row += 2
        ws_sum.cell(row=row, column=1, value="Disclaimer")
        ws_sum.cell(row=row, column=2, value=statement_meta.get("disclaimer", "This is a system-generated statement."))

        # Auto-size columns for "Summary"
        for col in ws_sum.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(val))
                except Exception:
                    pass
            ws_sum.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # --- Reorder sheets: Summary first, Invoices second ---
        wb._sheets = [ws_sum, ws_inv]

        # Write to BytesIO and return as response
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"statements_{start_date}_{end_date}.xlsx"
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )

    # 🔸 Export PDF (Flask HTML version, consistent with company statements)
    if export == "pdf":
        return render_template(
            'print_statement.html',
            start_date=start_date,
            end_date=end_date,
            total_invoices=total_invoices,
            total_amount=round(total_amount, 2),
            inv_rows=[
                {
                    "invoice_no": inv.invoiceId,
                    "date": inv.createdAt.strftime('%Y-%m-%d'),
                    "total": float(inv.totalAmount or 0),
                    "company": inv.customer.company if inv.customer else "(No Company)",
                    "phone": inv.customer.phone if inv.customer else "(No Phone)",
                }
                for inv in invs
            ],
            per_customer=per_customer,
            phone=phone,
            date_wise=True,
            APP_INFO=APP_INFO,
        )

    # 🔸 Default: HTML View
    return render_template(
        'statement.html',
        start_date=start_date,
        end_date=end_date,
        total_invoices=total_invoices,
        total_amount=total_amount,
        phone=phone,
        per_customer=per_customer,
        invs=invs,
        scope=scope,
    )


@app.route('/statements_company', methods=['GET', 'POST'])
def statements_company():
    """
    Generate HTML/CSV statements filtered by company name or phone number.
    Enhancements:
    - Unified date range parsing (with safe defaults)
    - Uses joinedload for performance
    - Adds graceful handling when no results or invalid input
    - Consistent CSV formatting with per-company totals
    - Improved suggestions list
    """

    # --- Query params ---
    query = (request.args.get('query') or '').strip()
    # Support both legacy 'fmt' and new 'format' param for export type
    fmt = (request.args.get('format') or request.args.get('fmt') or 'html').lower()
    start = request.args.get('start')
    end = request.args.get('end')
    phone = (request.args.get('phone') or '').strip()

    today = datetime.now(timezone.utc).date()
    start_dt = datetime(today.year, 1, 1, tzinfo=timezone.utc)
    end_dt = datetime.now(timezone.utc)

    # --- Parse date range if provided ---
    if start and end:
        try:
            start_dt = datetime.strptime(start, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            end_dt = datetime.strptime(end, '%Y-%m-%d').replace(tzinfo=timezone.utc) + timedelta(days=1) - timedelta(
                seconds=1)
            # 🔹 Enforce lower limit from info.json
            min_allowed_start = get_default_statement_start()
            if start_dt < min_allowed_start:
                start_dt = min_allowed_start
        except Exception:
            pass
    else:
        start_dt = get_default_statement_start()

    # --- Suggestions for autocomplete ---
    suggestions = []
    try:
        suggestions = [
            {"company": c.company or "", "phone": c.phone}
            for c in customer.query.filter(customer.isDeleted == False).all()
            if (c.company or c.phone)
        ]
    except Exception as e:
        print("[warn] Failed to load suggestions:", e)

    invs = []
    total_invoices = 0
    total_amount = 0.0

    # --- Query invoices if search provided ---
    if phone:
        # If phone is provided, search by exact phone
        q = (
            invoice.query
            .options(joinedload(invoice.customer))
            .join(customer, invoice.customerId == customer.id)
            .filter(
                invoice.isDeleted == False,
                customer.isDeleted == False,
                invoice.createdAt >= start_dt,
                invoice.createdAt <= end_dt,
                customer.phone == phone
            )
        )
        invs = q.order_by(invoice.createdAt.desc()).all()
        total_invoices = len(invs)
        total_amount = sum(float(inv.totalAmount or 0) for inv in invs)
    elif query:
        # Fallback: fuzzy search by company or phone (legacy)
        q = (
            invoice.query
            .options(joinedload(invoice.customer))
            .join(customer, invoice.customerId == customer.id)
            .filter(
                invoice.isDeleted == False,
                customer.isDeleted == False,
                invoice.createdAt >= start_dt,
                invoice.createdAt <= end_dt,
                or_(
                    func.lower(customer.company).like(f"%{query.lower()}%"),
                    customer.phone == query
                )
            )
        )
        invs = q.order_by(invoice.createdAt.desc()).all()
        total_invoices = len(invs)
        total_amount = sum(float(inv.totalAmount or 0) for inv in invs)

    # --- CSV Export ---
    if fmt == 'csv' and (phone or query):
        buf = io.StringIO()
        writer = csv.writer(buf)

        # Prepare header values
        customer_company = ''
        customer_phone = ''
        if invs and invs[0].customer:
            customer_company = invs[0].customer.company or ''
            customer_phone = invs[0].customer.phone or ''

        # Header
        writer.writerow([f"{APP_INFO['business']['name']} - Customer Statement"])
        writer.writerow(["Generated On", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow(["Customer Name", customer_company])
        writer.writerow(["Customer Phone", customer_phone])
        writer.writerow(["Period", f"{start_dt.strftime('%Y-%m-%d')} to {end_dt.strftime('%Y-%m-%d')}"])
        writer.writerow([])

        # Table (remove Company and Phone columns; they are shown in header)
        writer.writerow(["Invoice No", "Date", "Total (INR)"])
        for inv in invs:
            writer.writerow([
                inv.invoiceId,
                inv.createdAt.strftime('%Y-%m-%d'),
                f"{float(inv.totalAmount or 0):.2f}"
            ])

        # Summary
        writer.writerow([])
        writer.writerow(["Summary"])
        writer.writerow(["Total Invoices", total_invoices])
        writer.writerow(["Total Amount (INR)", f"{total_amount:.2f}"])
        writer.writerow([])

        # Payment Info (static for now)
        writer.writerow(["Payment Information"])
        writer.writerow(["Account Name", APP_INFO["bank"]["account_name"]])
        writer.writerow(["Bank", f"{APP_INFO['bank']['bank_name']}, {APP_INFO['bank']['branch']}"])
        writer.writerow(["Account Number", APP_INFO["bank"]["account_number"]])
        writer.writerow(["IFSC", APP_INFO["bank"]["ifsc"]])
        writer.writerow(["PhonePe/GPay", APP_INFO["bank"]["bhim"]])
        writer.writerow([])

        writer.writerow(["Disclaimer", "This is a system-generated statement. No signature required."])

        safe_company = (customer_company or "company").replace(" ", "_").replace("/", "_")
        filename = f"{safe_company}_{datetime.now().strftime('%Y-%m-%d')}_statement.csv"

        return Response(
            buf.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    # --- XLSX Export ---
    if fmt == 'xlsx' and (phone or query):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_inv = wb.create_sheet("Invoices")

        biz = APP_INFO.get("business", {})
        bank = APP_INFO.get("bank", {})
        statement_meta = APP_INFO.get("statement", {})

        # Prepare header values
        customer_company = ''
        customer_phone = ''
        if invs and invs[0].customer:
            customer_company = invs[0].customer.company or ''
            customer_phone = invs[0].customer.phone or ''

        # --- Summary Sheet ---
        bold = Font(bold=True)
        currency_fmt = u'INR #,##0.00'
        row = 1
        ws_sum.cell(row=row, column=1,
                    value=f"{biz.get('name', 'Business')} - {statement_meta.get('header_title', 'Customer Statement')}").font = bold
        row += 1
        ws_sum.cell(row=row, column=1, value="Generated On")
        ws_sum.cell(row=row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        row += 1
        ws_sum.cell(row=row, column=1, value="Customer Name")
        ws_sum.cell(row=row, column=2, value=customer_company)
        row += 1
        ws_sum.cell(row=row, column=1, value="Customer Phone")
        ws_sum.cell(row=row, column=2, value=customer_phone)
        row += 1
        ws_sum.cell(row=row, column=1, value="Period")
        ws_sum.cell(row=row, column=2, value=f"{start_dt.strftime('%Y-%m-%d')} to {end_dt.strftime('%Y-%m-%d')}")
        row += 2
        ws_sum.cell(row=row, column=1, value="Summary").font = bold
        row += 1
        ws_sum.cell(row=row, column=1, value="Total Invoices")
        ws_sum.cell(row=row, column=2, value=total_invoices)
        row += 1
        ws_sum.cell(row=row, column=1, value="Total Amount (INR)")
        amt_cell = ws_sum.cell(row=row, column=2, value=round(total_amount, 2))
        amt_cell.number_format = currency_fmt
        row += 2
        ws_sum.cell(row=row, column=1, value="Payment Information").font = bold
        row += 1
        ws_sum.cell(row=row, column=1, value="Account Name")
        ws_sum.cell(row=row, column=2, value=bank.get("account_name", ""))
        row += 1
        ws_sum.cell(row=row, column=1, value="Bank")
        ws_sum.cell(row=row, column=2, value=f"{bank.get('bank_name', '')}, {bank.get('branch', '')}")
        row += 1
        ws_sum.cell(row=row, column=1, value="Account Number")
        ws_sum.cell(row=row, column=2, value=bank.get("account_number", ""))
        row += 1
        ws_sum.cell(row=row, column=1, value="IFSC")
        ws_sum.cell(row=row, column=2, value=bank.get("ifsc", ""))
        row += 1
        ws_sum.cell(row=row, column=1, value="PhonePe/GPay")
        ws_sum.cell(row=row, column=2, value=biz.get("upi_id", ""))
        row += 2
        ws_sum.cell(row=row, column=1, value="Disclaimer")
        ws_sum.cell(row=row, column=2, value=statement_meta.get("disclaimer", "This is a system-generated statement."))

        # Auto-size columns for summary
        for col in ws_sum.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            ws_sum.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # --- Invoices Sheet ---
        headers = ["Invoice No", "Date", "Total (INR)"]
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_num, col_name in enumerate(headers, 1):
            c = ws_inv.cell(row=1, column=col_num, value=col_name)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align

        for row_num, inv in enumerate(invs, 2):
            ws_inv.cell(row=row_num, column=1, value=inv.invoiceId)
            ws_inv.cell(row=row_num, column=2, value=inv.createdAt.strftime('%Y-%m-%d'))
            amt_cell = ws_inv.cell(row=row_num, column=3, value=round(inv.totalAmount or 0, 2))
            amt_cell.number_format = currency_fmt
            amt_cell.alignment = Alignment(horizontal="right")

        for col in ws_inv.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            ws_inv.column_dimensions[col_letter].width = min(max_len + 2, 40)

        # Set sheet order
        wb._sheets = [ws_sum, ws_inv]

        # Output response
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_company = (customer_company or "company").replace(" ", "_").replace("/", "_")
        filename = f"{safe_company}_{datetime.now().strftime('%Y-%m-%d')}_statement.xlsx"
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    customer_company = ''
    customer_phone = ''
    if invs and invs[0].customer:
        customer_company = invs[0].customer.company or ''
        customer_phone = invs[0].customer.phone or ''

    # --- PDF Export ---
    if fmt == 'pdf' and (phone or query):
        # Use print-friendly HTML template for print preview
        return render_template(
            'print_statement.html',
            query=query,
            start_date=start_dt.date(),
            end_date=end_dt.date(),
            total_invoices=total_invoices,
            total_amount=round(total_amount, 2),
            inv_rows=[
                {
                    "invoice_no": inv.invoiceId,
                    "date": inv.createdAt.strftime('%Y-%m-%d'),
                    "total": float(inv.totalAmount or 0),
                }
                for inv in invs
            ],
            suggestions=suggestions,
            request=request,
            customer_company=customer_company,
            customer_phone=customer_phone,
            company_wise=True,
            APP_INFO=APP_INFO,
        )

    # --- Build rows for HTML template ---
    inv_rows = []
    for inv in invs:
        inv_rows.append({
            "invoice_no": inv.invoiceId,
            "date": inv.createdAt.strftime('%Y-%m-%d'),
            "total": float(inv.totalAmount or 0),
        })

    # --- Render HTML ---
    return render_template(
        'statements_company.html',
        query=query,
        start_date=start_dt.date(),
        end_date=end_dt.date(),
        total_invoices=total_invoices,
        total_amount=round(total_amount, 2),
        inv_rows=inv_rows,
        suggestions=suggestions,
        request=request,
        customer_company=customer_company,
        customer_phone=customer_phone,
    )


@app.route('/statements/accounting', methods=['GET'])
def accounting_statement():
    """
    Accounting-ledger statement with optional printable PDF view.
    """
    today = datetime.now(timezone.utc).date()
    min_start = get_default_statement_start().date()

    default_start = ACCOUNTING_STATEMENT_DEFAULT_START

    start_date = _parse_date(request.args.get('start')) or default_start
    end_date = _parse_date(request.args.get('end')) or today
    if start_date < min_start:
        start_date = min_start
    if end_date < start_date:
        end_date = start_date

    txn_type = (request.args.get('type') or 'all').lower()
    customer_query = (request.args.get('customer') or '').strip()
    export = (request.args.get('export') or 'html').lower()

    start_dt = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=timezone.utc)
    end_dt = datetime.combine(end_date, datetime.max.time()).replace(tzinfo=timezone.utc)

    context = _build_accounting_statement_context(
        start_dt,
        end_dt,
        start_date,
        end_date,
        txn_type_filter=txn_type,
        customer_query=customer_query,
    )

    suggestions = []
    try:
        suggestions = [
            {"company": c.company or "", "phone": c.phone}
            for c in customer.query.filter(customer.isDeleted == False).all()
            if (c.company or c.phone)
        ]
    except Exception as e:
        print("[warn] Failed to load accounting suggestions:", e)

    template_payload = dict(context, APP_INFO=APP_INFO)
    template_payload["suggestions"] = suggestions

    if export == 'pdf':
        return render_template('print_accounting_statement.html', **template_payload)

    return render_template('accounting_statement.html', **template_payload)


@app.route('/qr_code', methods=['GET', 'POST'])
def qr_code():
    upi_variants = _get_upi_variants()
    default_variant = upi_variants[0] if upi_variants else {
        'key': 'primary',
        'upi_id': APP_INFO.get('upi_info', {}).get('upi_id') or APP_INFO.get('business', {}).get('upi_id', ''),
        'upi_name': APP_INFO.get('upi_info', {}).get('upi_name') or APP_INFO.get('business', {}).get('upi_name', ''),
    }

    return render_template('QR_code.html',
                           upi_id=default_variant.get('upi_id', ''),
                           upi_name=default_variant.get('upi_name', ''),
                           selected_variant=default_variant.get('key', 'primary'),
                           upi_variants=upi_variants,
                           qr_image=False)


@app.route('/generate_qr', methods=['GET', 'POST'])
def generate_qr():
    source = request.form if request.method == 'POST' else request.args
    amount = source.get('amount')

    upi_info_defaults = APP_INFO.get('upi_info', {}) if isinstance(APP_INFO.get('upi_info'), dict) else {}
    business_defaults = APP_INFO.get('business', {}) if isinstance(APP_INFO.get('business'), dict) else {}

    selected_variant = _find_upi_variant(source.get('upi_variant'))
    if selected_variant:
        upi_id = selected_variant.get('upi_id') or upi_info_defaults.get('upi_id') or business_defaults.get('upi_id')
        upi_name = selected_variant.get('upi_name') or upi_info_defaults.get('upi_name') or business_defaults.get('upi_name') or business_defaults.get('owner')
    else:
        upi_id = source.get('upi_id') or upi_info_defaults.get('upi_id') or business_defaults.get('upi_id')
        upi_name = source.get('upi_name') or upi_info_defaults.get('upi_name') or business_defaults.get('upi_name') or business_defaults.get('owner')

    company_name = business_defaults.get('name') or APP_INFO['business']['name']

    api_url = f"{request.host_url.rstrip('/')}/api/generate_upi_qr"
    params = {"upi_id": upi_id, "amount": amount, "company_name": company_name}

    try:
        resp = requests.get(api_url, params=params, timeout=5)
        if resp.status_code == 200:
            data = resp.json()
            qr_svg_base64 = data.get('qr_svg_base64')
            upi_url = data.get('upi_url')
        else:
            qr_svg_base64 = None
            upi_url = None
    except Exception as e:
        print(f"[ERROR] failed to fetch QR: {e}")
        qr_svg_base64 = None
        upi_url = None

    qr_details = {
        'upi_id': upi_id,
        'company_name': company_name,
        'amount': amount
    }

    return render_template(
        'qr_display.html',
        upi_id=upi_id,
        qr_image=True,
        amount=amount,
        upi_name=upi_name,
        qr_code=qr_svg_base64,
        upi_url=upi_url,
        business_name=APP_INFO['business']['name'],
        amount_to_words=amount_to_words(amount)
    )


def load_supabase_config():
    try:
        url = APP_INFO['supabase']['url']
        key = APP_INFO['supabase']['key']
        last_updated = APP_INFO['supabase']['last_uploaded']
        return url, key, last_updated
    except Exception as e:
        print(f"Could not load supabase config: {e}")
        return None, None, None


def _update_supabase_last_uploaded(timestamp: str) -> None:
    info_path = ensure_info_json()
    try:
        with open(info_path, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
    except Exception:
        payload = {"data": {}}

    payload.setdefault("data", {})
    supabase_info = dict(payload["data"].get("supabase", {}))
    supabase_info["last_uploaded"] = timestamp
    payload["data"]["supabase"] = supabase_info

    try:
        with open(info_path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2)
    except Exception as exc:
        print(f"[warn] Failed to update Supabase sync metadata: {exc}")
        return

    refresh_info_json()


def _update_supabase_last_incremental(timestamp: str) -> None:
    info_path = ensure_info_json()
    try:
        with open(info_path, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
    except Exception:
        payload = {"data": {}}

    payload.setdefault("data", {})
    supabase_info = dict(payload["data"].get("supabase", {}))
    supabase_info["last_incremental_uploaded"] = timestamp
    payload["data"]["supabase"] = supabase_info

    try:
        with open(info_path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2)
    except Exception as exc:
        print(f"[warn] Failed to update Supabase incremental metadata: {exc}")
        return

    refresh_info_json()


def _resolve_external_backup_dir() -> Optional[Path]:
    """Return configured external backup directory, or None if unset."""
    location = (APP_INFO.get('file_location') or '').strip()
    if not location:
        return None
    try:
        return Path(location).expanduser()
    except Exception as exc:
        print(f"[warn] Invalid file_location '{location}': {exc}")
        return None


def _prune_backup_dir(directory: Path) -> None:
    """Keep only the newest BACKUP_RETENTION backup files in directory."""
    try:
        backups = sorted(
            directory.glob(f"{DATABASE_FILENAME}.*.bak"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
    except Exception as exc:
        print(f"[warn] Failed to enumerate external backups in {directory}: {exc}")
        return

    for old_backup in backups[BACKUP_RETENTION:]:
        try:
            old_backup.unlink()
        except Exception as exc:
            print(f"[warn] Failed to prune external backup {old_backup}: {exc}")


def _copy_backup_to_external(backup_source: Optional[Path] = None) -> Optional[Path]:
    """Copy the DB (or provided backup file) to configured external location."""
    destination_dir = _resolve_external_backup_dir()
    if not destination_dir:
        return None

    try:
        destination_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        print(f"[warn] Unable to prepare external backup directory {destination_dir}: {exc}")
        return None

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    source_path = backup_source if backup_source and backup_source.exists() else DB_PATH
    if backup_source and backup_source.exists() and backup_source.suffix == '.bak':
        target_name = backup_source.name
    else:
        target_name = f"{DATABASE_FILENAME}.{timestamp}.bak"

    target_path = destination_dir / target_name

    try:
        shutil.copy2(source_path, target_path)
    except Exception as exc:
        print(f"[warn] Failed to copy backup to {target_path}: {exc}")
        return None

    _prune_backup_dir(destination_dir)
    return target_path


def _create_db_backup() -> Path | None:
    backup_dir = DATA_DIR / BACKUP_DIRNAME
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    backup_path = backup_dir / f"{DATABASE_FILENAME}.{timestamp}.bak"
    try:
        shutil.copy2(DB_PATH, backup_path)
    except Exception as exc:
        print(f"[warn] Failed to create DB backup: {exc}")
        return None

    try:
        backups = sorted(backup_dir.glob(f"{DATABASE_FILENAME}.*.bak"), key=lambda p: p.stat().st_mtime, reverse=True)
        for old_backup in backups[BACKUP_RETENTION:]:
            try:
                old_backup.unlink()
            except Exception as cleanup_exc:
                print(f"[warn] Failed to prune old backup {old_backup}: {cleanup_exc}")
    except Exception as exc:
        print(f"[warn] Failed to prune backups: {exc}")

    external_path = _copy_backup_to_external(backup_path)
    if external_path:
        print(f"[info] External backup saved to {external_path}")

    return backup_path


def _format_sync_timestamp(raw_value):
    if not raw_value:
        return "Never"
    try:
        parsed = parser.parse(str(raw_value))
    except Exception:
        return str(raw_value)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    local_dt = parsed.astimezone(tz.tzlocal())
    date_part = local_dt.strftime("%d %b %Y")
    time_part = local_dt.strftime("%I:%M %p").lstrip('0')
    return f"{date_part} · {time_part}"


def _latest_backup_path() -> Path | None:
    backup_dir = DATA_DIR / BACKUP_DIRNAME
    if not backup_dir.exists():
        return None
    try:
        backups = sorted(
            backup_dir.glob(f"{DATABASE_FILENAME}.*.bak"),
            key=lambda p: p.stat().st_mtime,
            reverse=True
        )
    except Exception as exc:
        print(f"[warn] Failed to inspect backups: {exc}")
        return None
    return backups[0] if backups else None


def _ensure_recent_backup_on_shutdown() -> None:
    try:
        last_backup = _latest_backup_path()
        if last_backup is None:
            created = _create_db_backup()
            if created:
                print(f"[info] Shutdown backup created (no prior backups): {created}")
            return

        now_utc = datetime.now(timezone.utc)
        last_backup_time = datetime.fromtimestamp(last_backup.stat().st_mtime, tz=timezone.utc)
        age = now_utc - last_backup_time
        if age > timedelta(days=BACKUP_MAX_AGE_DAYS):
            created = _create_db_backup()
            if created:
                print(f"[info] Shutdown backup created (stale backup): {created}")
    except Exception as exc:
        print(f"[warn] Backup-on-shutdown failed: {exc}")


atexit.register(_ensure_recent_backup_on_shutdown)


def _check_supabase_connectivity(url: str, timeout: float = 5.0) -> tuple[bool, str]:
    health_url = f"{url.rstrip('/')}/auth/v1/health"
    try:
        response = requests.get(health_url, timeout=timeout)
    except requests.RequestException as exc:
        return False, str(exc)

    if response.status_code >= 500:
        return False, f"Supabase health check failed with status {response.status_code}"

    return True, ""


@app.post('/supabase_sync_all')
def supabase_sync_all():
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    def _respond(status: str, message: str, category: str = 'info', code: int = 200, **extra):
        if is_ajax:
            payload = {"status": status, "message": message}
            payload.update(extra)
            return jsonify(payload), code
        flash(message, category)
        return redirect(url_for('more'))

    url, key, _ = load_supabase_config()
    if not url or not key:
        return _respond('error', 'Supabase credentials are missing. Please update Account Settings.', 'danger', 400)

    reachable, error_message = _check_supabase_connectivity(url)
    if not reachable:
        return _respond('error', 'Cloud sync unavailable right now — check your internet connection and try again.', 'warning', 503, details=error_message)

    backup_path = _create_db_backup()

    if backup_path:
        print(f"[info] Created database backup at {backup_path}")

    try:
        result = upload_full_database(url, key)
    except SupabaseUploadError as exc:
        skipped = ', '.join(exc.skipped_tables) if exc.skipped_tables else 'none'
        detail = f" Reason: {exc.detail}" if exc.detail else ''
        message = (
            f"Supabase rejected {exc.failed_count} records in '{exc.failed_table}'. "
            f"Upload stopped before tables: {skipped}.{detail}"
        )
        return _respond('error', message, 'danger', 500, details=message)
    except Exception as exc:
        return _respond('error', f'Failed to sync with Supabase: {exc}', 'danger', 500)

    timestamp = datetime.now(timezone.utc).isoformat()
    _update_supabase_last_uploaded(timestamp)

    success_message = f"Uploaded {result.uploaded} records to Supabase ({result.failed} failed)."
    return _respond('success', success_message, 'success', 200, uploaded=result.uploaded, failed=result.failed)


@app.post('/backup/local_copy')
def create_local_backup_copy():
    refresh_info_json()
    destination = _resolve_external_backup_dir()
    if not destination:
        flash('Set a backup folder in Account Settings before creating local copies.', 'warning')
        return redirect(url_for('more'))

    backup_path = _copy_backup_to_external()
    if backup_path:
        flash(f'Backup copied to {backup_path}', 'success')
    else:
        flash('Unable to copy database to the configured folder. Check permissions and path.', 'danger')
    return redirect(url_for('more'))


@app.post('/backup/snapshot')
def create_backup_snapshot():
    backup_path = _create_db_backup()
    if backup_path:
        flash(f'Backup snapshot created at {backup_path}', 'success')
    else:
        flash('Unable to create backup snapshot. Check disk space and permissions.', 'danger')
    return redirect(url_for('more'))


@app.post('/supabase_sync_incremental')
def supabase_sync_incremental():
    url, key, _ = load_supabase_config()
    if not url or not key:
        return jsonify({
            "status": "error",
            "message": "Supabase credentials are missing. Please update Account Settings."
        }), 400

    reachable, error_message = _check_supabase_connectivity(url)
    if not reachable:
        return jsonify({
            "status": "error",
            "message": "Looks like the internet is down. Please reconnect and try again."
        }), 503

    try:
        result = upload_to_supabase(url, key)
    except Exception as exc:
        return jsonify({
            "status": "error",
            "message": f"Failed to upload incremental data: {exc}"
        }), 500

    db_result = result["db"]
    analytics_result = result["analytics"]
    archived = result["archived"]

    payload = {
        "status": "ok",
        "db_uploaded": db_result.uploaded,
        "db_failed": db_result.failed,
        "analytics_uploaded": analytics_result.uploaded,
        "analytics_failed": analytics_result.failed,
        "archived": archived,
    }

    if db_result.failed == 0 and analytics_result.failed == 0:
        timestamp = datetime.now(timezone.utc).isoformat()
        _update_supabase_last_incremental(timestamp)
        payload["message"] = (
            f"Uploaded {db_result.uploaded} changes"
            f" and {analytics_result.uploaded} analytics events."
        )
        payload["last_uploaded"] = _format_sync_timestamp(timestamp)
    else:
        payload["status"] = "partial"
        payload["message"] = (
            "Incremental sync finished with errors. "
            "Check logs/failed for details."
        )
        supabase_meta = APP_INFO.get('supabase', {})
        previous = supabase_meta.get('last_incremental_uploaded') or supabase_meta.get('last_uploaded')
        payload["last_uploaded"] = _format_sync_timestamp(previous)

    return jsonify(payload)


if __name__ == '__main__':
    host = os.getenv('HOST', '0.0.0.0')
    port = int(os.getenv('PORT', 5000))

    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

    try:
        sock.bind((host, port))
        sock.close()
    except OSError as exc:
        port = random.randint(5001, 5999)
        print(f"Port 5000 busy, switching to port {port}")

    print(f"Starting WSGI server on https://{host}:{port}")
    serve(app, host=host, port=port)
